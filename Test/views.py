import json
import pydoc
import re
from django.apps import apps
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render,redirect,get_object_or_404
from django.contrib.auth import authenticate, login ,logout,get_user_model
from Account.forms import RegistrationForm
from Account.models import *
from Form.models import Form
from Form.views import common_module_master
from Masters.models import *
import Db 
import bcrypt
from django.contrib.auth.decorators import login_required
from THRMS.encryption import *
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph
from Account.utils import decrypt_email, encrypt_email
import requests
import traceback
import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.conf import settings
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Border, Side
import calendar
from datetime import datetime, timedelta
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import AccessToken
from django.utils import timezone
from Account.models import *
from Masters.models import *
from Account.db_utils import callproc
from django.views.decorators.csrf import csrf_exempt
import os
from django.urls import reverse
from THRMS.settings import *
import logging
from django.http import FileResponse, Http404
import mimetypes

from Workflow.models import workflow_action_master, workflow_matrix
logger = logging.getLogger(__name__)

from Test.models import *
import random
from random import shuffle
from THRMS.encryption import enc, dec
from django.utils.timezone import now
from django.db.models import OuterRef, Subquery

from django.db.models.functions import TruncDate
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime
import io
from openpyxl.cell.cell import Cell
import platform

@login_required
def enterthedetails(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user = request.user.id 
        if request.method == "GET":
            ErrorMessage =request.GET.get('ErrorMessage', '')
            SuccessMessage =request.GET.get('SuccessMessage', '')            
            dpl = PostMaster.objects.values_list('id', 'post')
        if request.method == "POST":
            name1 = request.POST.get("text_name", "")
            email1 = request.POST.get("email", "")
            mobile1 = request.POST.get("mobile", "")
            post1 = request.POST.get("post", "")
            post_id = int(post1)
            
            status1 = 2 if post_id in [4, 7] else 1           
                                                          
            # noq1 = request.POST.get("noq", "")
            # noq = int(noq1)
            # Get and shuffle questions
            # all_questions = list(QuestionAnswerMaster.objects.filter(post=post1))
            # if not all_questions:
            #     questions_var = "no_questions"
            # else:
            #     questions_var = "questions_exist" 
                
            # if all_questions:                
            #     shuffle(all_questions)
            #     selected_questions = all_questions[:min(noq, len(all_questions))]

            #     # Extract question IDs
            #     question_ids = [str(q.question_id) for q in selected_questions]
            #     ids_str1 = ",".join(question_ids)
            #     ids_str = enc(str(ids_str1))

            CandidateTestMaster.objects.create(
                name=name1,
                email=email1,
                mobile=mobile1,
                post=post1,
                created_at=timezone.now(),
                created_by=user,
                status=status1
            )  
                                    
                # candidate_id2 = (
                #     CandidateTestMaster.objects
                #     .filter(name=name1, mobile=mobile1)
                #     .order_by('-id')  # or use '-created_at' if available
                #     .values_list('id', flat=True)
                #     .first()
                # )
                
                # candidate_id = enc(str(candidate_id2))                
                                                                
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])
        messages.error(request,"Some Error Occurred !!") 
        return redirect('enterthedetails')                  
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            return render(request, "Test/enter_details.html", {'dpl': dpl})
        elif request.method == "POST":
            messages.success(request,"Details Submitted Successfully !") 
            return redirect('candidate_index')            
            # if questions_var == "no_questions":
            #     messages.error(request,"No Questions Found For This Post !!!!") 
            #     return redirect('candidate_index')
            # else:                
            #     return redirect(f'/test_page?zparqwwq={ids_str}&candidate_id={candidate_id}')                     
        
# @login_required
# def test_page(request):
#     Db.closeConnection()
#     m = Db.get_connection()
#     cursor = m.cursor()
    
#     try:
#         user = request.user.id 
#         if request.method == "GET":
#             start_time1 = timezone.now()
#             start_time = enc(str(start_time1))
#             candidate_id1 = request.GET.get("candidate_id", "")
#             # candidate_id=dec(candidate_id1)
#             ids_str1 = request.GET.get("zparqwwq", "") #this are question ids 
#             ids_str=dec(ids_str1)
#             id_list = [int(qid) for qid in ids_str.split(",") if qid.isdigit()]  
            
#             questions_raw = QuestionAnswerMaster.objects.filter(question_id__in=id_list)
#             questions_raw = sorted(questions_raw, key=lambda q: id_list.index(q.question_id))  
            
#             questions = []
#             for q in questions_raw:
#                 options = [
#                     {"text": q.choice1},
#                     {"text": q.choice2},
#                     {"text": q.choice3},
#                     {"text": q.choice4},
#                 ]
#                 random.shuffle(options)  # shuffle in-place

#                 questions.append({
#                     "question_id": q.question_id,
#                     "question": q.question,
#                     "options": options,
#                 })                     

#         if request.method == "POST":

#             question_ids = request.POST.getlist("question_ids")
#             candidate_id1 = request.POST.get("candidate_id", "")
#             candidate_id=dec(candidate_id1)
                         
#             start_time1 = request.POST.get("start_time", "")
#             start_time_str=dec(start_time1)             

            
# # GET CANDIDATE DETAILS            
#             candidate_data = CandidateTestMaster.objects.filter(id=candidate_id).values('name', 'post').first()
#             candidate_name = candidate_data['name']
#             post1 = candidate_data['post']
#             start_time = enc(str(start_time1))


# # TIME TAKEN BY CANDIDATE
            
#             start_time = datetime.fromisoformat(start_time_str)
#             end_time = now()
#             time_taken = end_time - start_time
#             # Convert to minutes and seconds
#             seconds_taken = int(time_taken.total_seconds())
#             hours = seconds_taken // 3600
#             minutes = (seconds_taken % 3600) // 60
#             seconds = seconds_taken % 60

#             # Optional: format as string "mm:ss"
#             time_taken_str = f"{hours:02}:{minutes:02}:{seconds:02}"  # HH:MM:SS           

# # SCORE CALCULATION OF CANDIDATE 
#             score = 0
#             total = len(question_ids)
#             result_details = []            
#             for qid in question_ids:
#                 question = QuestionAnswerMaster.objects.get(pk=int(qid))
#                 user_answer = request.POST.get(f"answer_{qid}", "")
#                 is_correct = (user_answer == question.correct_answer) 

#                 if is_correct:
#                     isright1 = "Yes"
#                     score += 1 
#                 else:
#                     isright1 = "No"     
                    
#                 candidate_answer = CandidateAnswer(
#                     candidate_id=candidate_id,
#                     post=post1,
#                     question_id=qid,
#                     candidates_answer=user_answer,
#                     is_right=isright1,
#                     created_at=timezone.now(),
#                     created_by=user
#                 )

#                 candidate_answer.save()
                
# # PERCENTAGE CALCULATION & STATUS
#             percentage = (score / total) * 100 
#             if percentage >= 40:
#                 status = "pass"
#             else:
#                 status = "fail"   
                
#             CandidateTestMaster.objects.filter(id=candidate_id).update(
#                 marks_received=score,
#                 out_of=total,
#                 time_taken=time_taken_str,
#                 percentage=percentage,
#                 status=status,
#                 updated_by=user,
#                 updated_at=timezone.now(),
#                 test_start_time=start_time,
#                 test_end_time=end_time
#             )                                                                                                       
            
#     except Exception as e:
#         print("error-" + e)
#         tb = traceback.extract_tb(e.__traceback__)
#         fun = tb[0].name
#         cursor.callproc("stp_error_log",[fun,str(e),request.user.id])          
#     finally:
#         cursor.close()
#         m.commit()
#         m.close()
#         Db.closeConnection()
#         if request.method == "GET":
#             return render(request, "Test/test_page.html",{"questions": questions,"candidate_id":candidate_id1,"start_time":start_time})
#         elif request.method == "POST":
#             return redirect(f'/result_page?cname={candidate_id1}')     
        
@login_required
def result_page(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        if request.method == "GET":                                    
            candidate_id1 = request.GET.get("cname", "")
            typeofscreen = request.GET.get("type", "")
            if not typeofscreen:
                typeofscreen=None                              
            candidate_id=dec(candidate_id1)       
            # Fetch the candidate record by ID and get specific fields
            candidate = CandidateTestMaster.objects.filter(candidate_id=candidate_id).values('name', 'time_taken', 'percentage', 'performance_level', 'it_percentage', 'it_performance_level').first()

            # Initialize variables to None in case no record is found
            name = time_taken = percentage = status = None

            # If candidate record exists, extract fields into separate variables
            if candidate:
                name = candidate['name']
                time_taken = candidate['time_taken']
                percentage = candidate['percentage']
                performance_level = candidate['performance_level']
                it_percentage = candidate['it_percentage'] 
                it_performance_level = candidate['it_performance_level']                                                              
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])          
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            return render(request, "Test/result_page.html",{"candidate_name":name,"time_taken":time_taken,"percentage":percentage,"performance_level":performance_level,"it_percentage":it_percentage,"it_performance_level":it_performance_level,"typeofscreen":typeofscreen})
        
@login_required
def test_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        if request.method == "GET": 
            
            dpl = PostMaster.objects.values_list('id', 'post')            
            status_list = CandidateTestMaster.objects.values_list('status', 'status').distinct()
            performance_level_list = PerformanceMaster.objects.values_list('performance_level', 'performance_level').distinct()
            
            cursor.callproc("stp_getCreatedAt") # all the types from parameter master
            for result in cursor.stored_results():
                created_at = list(result.fetchall())                        
                        
            # Subquery to fetch post name from PostMaster
            post_name_subquery = PostMaster.objects.filter(
                id=OuterRef('post')  # both are IntegerFields now
            ).values('post')[:1]

            # Annotate queryset with post name
            queryset = CandidateTestMaster.objects.annotate(
                post_name=Subquery(post_name_subquery)
            )            
                                               
            candidate_data = []
            for item in queryset:
                row = {
                    'id': item.id,
                    'name': item.name,
                    'mobile':item.mobile,
                    'email': item.email,
                    'post': item.post_name,                    
                    'status': item.status,
                    'percentage': item.percentage,
                    'time_taken': item.time_taken,
                    'status': item.status,
                    'created_at': item.created_at,
                    'Encryp': enc(str(item.id)),                    
                }
                candidate_data.append(row)                     
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])          
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            return render(request, "Test/test_index.html",{'data': candidate_data,'dpl':dpl,'status_list':status_list,'created_at':created_at,'performance_level_list':performance_level_list})

@login_required
def partial_details_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()    
    try:
        if request.method == "POST":
            post_id = request.POST.get("dp", "") or None
            skillset = request.POST.get("skillset", "") or None
            it_performance = request.POST.get("it_performance", "") or None            
            created_at = request.POST.get("created_at", "") or None
            if created_at:
                # Convert '23-05-2025' to '2025-05-23'
                created_date = datetime.strptime(created_at, "%d-%m-%Y").date()
                created_date_str = created_date.strftime('%Y-%m-%d') 
            else:
                created_date_str = None                           
         
            param = [post_id or None, skillset or None, created_date_str or None,it_performance or None]
            candidate_details = []
            cursor.callproc("stp_getTestIndex", param)
            for result in cursor.stored_results():
                candidate_details  = list(result.fetchall())             

            candidate_data = [
                {
                    'id': row[0],
                    'name': row[1],
                    'mobile': row[2],
                    'email': row[3],
                    'post': row[4],
                    'status': row[5],
                    'percentage': row[6],
                    'time_taken': row[7],
                    'created_at': row[8],
                    'performance_level': row[9],
                    'it_performance_level': row[10],                    
                    'Encryp': enc(str(row[0])),
                    'enc_candidate_id': enc(str(row[11])),
                    'enc_form_data_id': enc(str(row[12])),                    
                }
                for row in candidate_details
            ]

            # Render the HTML partial
            table = render_to_string(
                "Test/partial_test_index.html",  # replace with your actual template path
                {
                    "result_set": candidate_data,
                    "displayIndex": 1,
                },
            )

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id]) 
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "POST":
            return JsonResponse({"table": table}, safe=False)
        
@login_required
def partial_details_index_onpageload(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()    
    try:
        if request.method == "POST":

            candidate_details = []
            cursor.callproc("stp_getTestIndex_onpageload")
            for result in cursor.stored_results():
                candidate_details  = list(result.fetchall())             

            candidate_data = [
                {
                    'id': row[0],
                    'name': row[1],
                    'mobile': row[2],
                    'email': row[3],
                    'post': row[4],
                    'status': row[5],
                    'percentage': row[6],
                    'time_taken': row[7],
                    'created_at': row[8],
                    'performance_level': row[9],
                    'it_performance_level': row[10],
                    'Encryp': enc(str(row[0])),
                    'enc_candidate_id': enc(str(row[11])),
                    'enc_form_data_id': enc(str(row[12])),                    
                }
                for row in candidate_details
            ]

            # Render the HTML partial
            table = render_to_string(
                "Test/partial_test_index.html",  # replace with your actual template path
                {
                    "result_set": candidate_data,
                    "displayIndex": 1,
                },
            )

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id]) 
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "POST":
            return JsonResponse({"table": table}, safe=False)

@login_required        
def download_candidate_excel(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()      
    try:
        if request.method == "POST":
            post_id = request.POST.get("post", "") or None
            skillset = request.POST.get("skillset", "") or None
            it_performance = request.POST.get("it_performance", "") or None
            created_at = request.POST.get("created_at", "") or None

            if created_at:
                # Convert '23-05-2025' to '2025-05-23'
                created_date = datetime.strptime(created_at, "%d-%m-%Y").date()
                created_date_str = created_date.strftime('%Y-%m-%d')
            else:
                created_date_str = None

            param = [post_id or None, skillset or None, created_date_str or None, it_performance or None]
            candidate_details = []

            cursor.callproc("stp_getTestIndexForExcel", param)
            for result in cursor.stored_results():
                candidate_details = list(result.fetchall())

            # Prepare Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Candidate Test Details"

            # Styles
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Merge cells from A1 to H1 for title row
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

            # Set title text and style
            title_cell = ws.cell(row=1, column=1, value="CANDIDATE DETAILS")
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = center_align

            # Apply thin border to all cells in the merged range A1:H1
            for col in range(1, 9):  # Columns A to H
                cell = ws.cell(row=1, column=col)
                cell.border = thin_border

            # Header Row
            headers = ['Name', 'Mobile', 'Email', 'Post', 'Time Taken','Skillset', 'IT Performance', 'Created At']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col_num, value=header)
                cell.font = bold_font
                cell.alignment = center_align
                cell.border = thin_border

            for row_num, row_data in enumerate(candidate_details, start=3):
                data_without_id = row_data[1:]  # remove the first element (ID)
                for col_num, cell_value in enumerate(data_without_id, 1):
                    # Format "Created At" (now at 8th position)
                    if col_num == 8 and isinstance(cell_value, datetime):
                        if platform.system() == 'Windows':
                            time_str = cell_value.strftime('%B %d, %Y, %#I:%M %p')
                        else:
                            time_str = cell_value.strftime('%B %d, %Y, %-I:%M %p')
                        cell_value = time_str.replace('AM', 'a.m.').replace('PM', 'p.m.')

                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.alignment = center_align
                    cell.border = thin_border

            # Adjust column width

            for col in ws.columns:
                max_length = 0
                column_letter = None

                for cell in col:
                    if isinstance(cell, Cell):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))

                if column_letter:
                    ws.column_dimensions[column_letter].width = max_length + 2

            # Save to in-memory stream
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Prepare HTTP response
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="candidate_test_details.xlsx"'
            return response
               
    except Exception as e:
        try:
            tb = traceback.extract_tb(e.__traceback__)
            fun = tb[0].name
            cursor.callproc("stp_error_log", [fun, str(e), request.user.id])
        except:
            pass  # Optional: prevent infinite error loop if error logging fails
        return HttpResponse("An error occurred while generating the Excel file.", status=500)
    finally : 
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        
@login_required
def candidate_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    role = str(request.session.get('role_id'))
    user= request.session.get('user_id')
    
    try:
        if request.method == "GET":
            if request.user.is_authenticated ==True:                
                user = request.user.id              
                TemporaryQuestion.objects.filter(user_id=user).delete()
                                            
            # Subquery to fetch post name from PostMaster
            post_name_subquery = PostMaster.objects.filter(
                id=OuterRef('post') 
            ).values('post')[:1]
            table_data = []

            workflows = workflow_matrix.objects.all()

            # Filter in Python since DB collation doesn't allow regex
            workflow = None
            for wf in workflows:
                role_ids = [r.strip() for r in wf.role_id.split(',')]
                if role in role_ids:
                    workflow = wf
                    break
            form_id = workflow.form_id
            action_id = workflow.button_type_id

            form = get_object_or_404(Form, id = form_id )

            module_id = form.module
            module_tables = common_module_master(module_id)

            IndexTable = apps.get_model('Form', module_tables["index_table"])
            DataTable = apps.get_model('Form', module_tables["data_table"])
            FileTable = apps.get_model('Form', module_tables["file_table"])

            status = get_object_or_404(IndexTable, created_by = user).status

            form_data = IndexTable.objects.filter(created_by=user)
            
            created_ids = set(form_data.values_list('id', flat=True))  # Get all matching IndexTable IDs

            forms = callproc("stp_get_forms", ['view_form'])  
            sf = 'view_form_Candidate_Form_I' 
            header = callproc("stp_get_view_form_header", [sf])          
            rows = callproc("stp_get_view_forms", [sf]) 

            table_data = []

            for row in rows:
                desc_id = row[0]  # First column is assumed to be the IndexTable ID
                if desc_id in created_ids:  # Only include rows where the user is the creator
                    encrypted_id = enc(str(desc_id))
                    table_data.append((encrypted_id,) + row[1:])


            # Annotate queryset with post name
            # queryset = CandidateTestMaster.objects.annotate(
            #     post_name=Subquery(post_name_subquery)
            # )            

            # Get the latest candidate record (by ID) for created_by=24 and status=2
            queryset = CandidateTestMaster.objects.filter(
                created_by=user,
                status='4'  # Use 2 if it's an IntegerField, '2' if it's CharField
            ).order_by('-id').annotate(
                post_name=Subquery(post_name_subquery)
            )[:1]  # Limit to 1 record
                                               
            candidate_data = []
            last_status=0
            for item in queryset:
                row = {
                    'id': item.id,
                    'name': item.name,
                    'mobile':item.mobile,
                    'email': item.email,
                    'post': item.post_name,
                    'status': item.status,
                    'percentage': item.percentage,
                    'time_taken': item.time_taken,
                    'status': item.status,
                    'created_at': item.created_at,
                    'Encryp': enc(str(item.id)),
                    'created_at': item.created_at,                                        
                }
                candidate_data.append(row) 
                last_status = item.status                        
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])          
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            return render(request, "Test/candidate_index.html",{'data': candidate_data,'last_status':last_status,'forms':forms,'rows':rows,'header':header,'table_data':table_data,'status':status})
        
@login_required
def test_page(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user = request.user.id 
        if request.method == "GET":
            redir_path=None
            start_time1 = timezone.now()
            start_time = enc(str(start_time1))
            form_data_id1 = request.GET.get("id", "")
            form_data_id=dec(form_data_id1)
            
            parameter_for = [form_data_id]
            cursor.callproc("stp_getPostOfCandidate",parameter_for)
            for result in cursor.stored_results():
                post_list = result.fetchall()
                post_id = post_list[3][0]  
                
            po_id = post_id
            
            cursor.callproc("stp_getCandidateId",parameter_for)
            for result3 in cursor.stored_results():
                post_list3 = result3.fetchall()
                candidate_id = post_list3[0][0]  # Index 0 record
                
            candidate_id_enc = enc(str(candidate_id))                             
                        
            # candidate = CandidateTestMaster.objects.get(id=candidate_id)
            # po_id = candidate.post          
            
            existing = TemporaryQuestion.objects.filter(candidate_id=candidate_id, status=1).exists()            
            if not existing:            
                param_post=[po_id]
                cursor.callproc("stp_getPostTopicQuestions", param_post)
                for result in cursor.stored_results():
                    post_topic_questions = list(result.fetchall())  # returns tuples            
                if post_topic_questions:
                    topic_question_map = {}
                    for row in post_topic_questions:
                        _, topic_id, question_id, no_of_questions = row
                        if topic_id not in topic_question_map:
                            topic_question_map[topic_id] = {"limit": no_of_questions, "questions": []}
                        topic_question_map[topic_id]["questions"].append(question_id)
                        
                    final_question_ids = []

                    for topic_id, data in topic_question_map.items():
                        q_ids = data["questions"]
                        shuffle(q_ids)
                        final_question_ids.extend(q_ids[:data["limit"]])                

                    shuffle(final_question_ids)

                    csv_question_ids = ",".join(map(str, final_question_ids))

                    # Create entry in temporary_questions table
                    if not TemporaryQuestion.objects.filter(candidate_id=candidate_id, status=1).exists():            
                        TemporaryQuestion.objects.create(
                            user_id=request.user.id, 
                            candidate_id=candidate_id,  # or request.user.id
                            question_ids=csv_question_ids,
                            status=1,
                            created_by=request.user.id,  # same or admin/creator ID
                            created_at=datetime.now()
                        )
                    temp_q = TemporaryQuestion.objects.get(candidate_id=candidate_id, status=1)
                    final_question_ids = list(map(int, temp_q.question_ids.split(',')))
                else:
                    redir_path='nodata'                    
                    return redirect(f'/candidate_index')                      
            else:
                temp_q = TemporaryQuestion.objects.get(candidate_id=candidate_id, status=1)
                final_question_ids = list(map(int, temp_q.question_ids.split(',')))                    
            
            questions_raw = list(QuestionAnswerMaster.objects.filter(question_id__in=final_question_ids))
            # Maintain original order
            questions_raw = sorted(questions_raw, key=lambda q: final_question_ids.index(q.question_id))

            questions = []
            for q in questions_raw:
                options = [
                    {"text": q.choice1},
                    {"text": q.choice2},
                    {"text": q.choice3},
                    {"text": q.choice4},
                ]
                shuffle(options)

                questions.append({
                    "question_id": q.question_id,
                    "question": q.question,
                    "options": options,
                })                  

        if request.method == "POST":

            question_ids = request.POST.getlist("question_ids")
            candidate_id1 = request.POST.get("candidate_id", "")
            candidate_id=dec(candidate_id1)
            form_data_id1 = request.POST.get("form_data_id", "")
            form_data_id=dec(form_data_id1)
                                                 
            start_time1 = request.POST.get("start_time", "")
            start_time_str=dec(start_time1)             

            
# GET CANDIDATE DETAILS            
            # candidate_data = CandidateTestMaster.objects.filter(id=candidate_id).values('name', 'post').first()
            # candidate_name = candidate_data['name']
            
            parameter_for_2 = [form_data_id]
            cursor.callproc("stp_getPostOfCandidate",parameter_for_2)
            for result2 in cursor.stored_results():
                post_list1 = result2.fetchall()
                nme = post_list1[0][0]  # Index 0 record
                eml = post_list1[1][0]
                mbno = post_list1[2][0]
                post_id1 = post_list1[3][0]                                                            
            
            post1 = post_id1
            start_time = enc(str(start_time1))


# TIME TAKEN BY CANDIDATE
            
            start_time = datetime.fromisoformat(start_time_str)
            end_time = now()
            time_taken = end_time - start_time
            # Convert to minutes and seconds
            seconds_taken = int(time_taken.total_seconds())
            hours = seconds_taken // 3600
            minutes = (seconds_taken % 3600) // 60
            seconds = seconds_taken % 60

            # Optional: format as string "mm:ss"
            time_taken_str = f"{hours:02}:{minutes:02}:{seconds:02}"  # HH:MM:SS           

# SCORE CALCULATION OF CANDIDATE 
            it_topic_id = 5
            it_score = 0
            it_total = 0
            other_score = 0
            other_total = 0       

            for qid in question_ids:
                try:
                    question = QuestionAnswerMaster.objects.get(pk=int(qid))
                except QuestionAnswerMaster.DoesNotExist:
                    continue

                user_answer = request.POST.get(f"answer_{qid}", "")
                is_correct = (user_answer == question.correct_answer)

                isright1 = "Yes" if is_correct else "No"

                # Update IT / Other scores
                if question.topic_id == it_topic_id:
                    it_total += 1
                    if is_correct:
                        it_score += 1
                else:
                    other_total += 1
                    if is_correct:
                        other_score += 1            
                    
                candidate_answer = CandidateAnswer(
                    candidate_id=candidate_id,
                    post=post1,
                    question_id=qid,
                    candidates_answer=user_answer,
                    is_right=isright1,
                    created_at=timezone.now(),
                    created_by=user
                )

                candidate_answer.save()
                
# PERCENTAGE CALCULATION & STATUS
            # Compute percentages
            it_percentage = round((it_score / it_total) * 100, 2) if it_total else 0
            other_percentage = round((other_score / other_total) * 100, 2) if other_total else 0

            # Get performance levels
            def get_performance(percentage):
                perf = PerformanceMaster.objects.filter(
                    percentage_from__lte=percentage,
                    percentage_upto__gte=percentage
                ).first()
                return perf.performance_level if perf else "Unknown"

            it_performance = get_performance(it_percentage)
            other_performance = get_performance(other_percentage)

            CandidateTestMaster.objects.create(
                candidate_id=candidate_id,
                it_marks_received=it_score,
                it_out_of=it_total,
                it_percentage=it_percentage,
                it_performance_level=it_performance,
                marks_received=other_score,
                out_of=other_total,
                percentage=other_percentage,
                performance_level=other_performance,
                time_taken=time_taken_str,
                status=2,
                test_start_time=start_time,
                test_end_time=end_time,
                created_by=user,
                created_at=timezone.now(),
                name=nme,
                mobile=mbno,
                email=eml,
                form_data_id=form_data_id,
                post=post1
            )

            TemporaryQuestion.objects.filter(candidate_id=candidate_id).delete()                                                                                                     
            
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])          
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            if redir_path == "nodata":
                messages.warning(request,"Questions Not Assigned for the Post !!")                
                return redirect(f'/candidate_index')            
            else:    
                return render(request, "Test/test_page.html",{"questions": questions,"candidate_id":candidate_id_enc,"start_time":start_time,"form_data_id":form_data_id1})
        elif request.method == "POST":
            return redirect(f'/result_page?cname={candidate_id1}')
        
@login_required
def edit_candidate(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user = request.user.id 
        if request.method == "GET":
            ErrorMessage =request.GET.get('ErrorMessage', '')
            SuccessMessage =request.GET.get('SuccessMessage', '')
            candidate_id1 = request.GET.get("id", "")
            candidate_id=dec(candidate_id1)                                    
            dpl = PostMaster.objects.values_list('id', 'post')
            candidate_data = CandidateTestMaster.objects.filter(id=candidate_id).values('name', 'email', 'mobile', 'post').first()

            context = {
                "name": candidate_data['name'] if candidate_data else None,
                "email": candidate_data['email'] if candidate_data else None,
                "mobile": candidate_data['mobile'] if candidate_data else None,
                "post": candidate_data['post'] if candidate_data else None,
            }
            context['dpl'] = dpl
            context['candidate_id1'] = candidate_id1
                        
        if request.method == "POST":
            name1 = request.POST.get("text_name", "")
            email1 = request.POST.get("email", "")
            mobile1 = request.POST.get("mobile", "")
            post1 = request.POST.get("post", "")
            post_id = int(post1)
            candidate_id2 = request.POST.get("candidate_id", "")
            candidate_idd=dec(candidate_id2)            
                     
            CandidateTestMaster.objects.filter(id=candidate_idd).update(
                name=name1,
                email=email1,
                mobile=mobile1,
                post=post1,
                updated_at=timezone.now(),
                updated_by=user,
            ) 
                                                                                
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])
        messages.error(request,"Some Error Occurred !!") 
        return redirect('candidate_index')                  
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method == "GET":
            return render(request, "Test/edit_candidate.html", context)
        elif request.method == "POST":
            messages.success(request,"Details Submitted Successfully !") 
            return redirect('candidate_index')

@login_required
def view_candidate_details(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    
    try:
        user = request.user.id 
        if request.method == "GET":
            ErrorMessage =request.GET.get('ErrorMessage', '')
            SuccessMessage =request.GET.get('SuccessMessage', '')
            candidate_id1 = request.GET.get("id", "")
            candidate_id=dec(candidate_id1)                                    
            dpl = PostMaster.objects.values_list('id', 'post')
            # candidate_data = CandidateTestMaster.objects.filter(id=candidate_id).values('name', 'email', 'mobile', 'post').first()

            candidate_data = CandidateTestMaster.objects.filter(id=candidate_id).values(
                'name', 'email', 'mobile', 'post'
            ).first()

            # Default post name
            post_name = None

            # If candidate data exists and post is available, get post name
            if candidate_data and candidate_data.get('post'):
                post_obj = PostMaster.objects.filter(id=candidate_data['post']).first()
                if post_obj:
                    post_name = post_obj.post

            context = {
                "name": candidate_data['name'] if candidate_data else None,
                "email": candidate_data['email'] if candidate_data else None,
                "mobile": candidate_data['mobile'] if candidate_data else None,
                "post": candidate_data['post'] if candidate_data else None,
                "post_name": post_name,
                "dpl": dpl,
                "candidate_id1": candidate_id1,
            }
                                                                                                        
    except Exception as e:
        print("error-" + e)
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),request.user.id])
        messages.error(request,"Some Error Occurred !!") 
        return redirect('view_candidate_details')                  
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        return render(request, "Test/view_candidate_details.html", context)

@login_required
def view_candidate_answersheet(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    try:
        if request.method == "GET":        
            candidate_id1 = request.GET.get("id", "")
            candidate_id = dec(candidate_id1)  # decrypt if needed
            candidate_name = CandidateTestMaster.objects.get(candidate_id=candidate_id).name                        
            para=[candidate_id]
            
            cursor.callproc("stp_getCandidateAnswer", para)
            for result in cursor.stored_results():
                raw_data = list(result.fetchall())

            questions = []
            for row in raw_data:
                q_dict = {
                    "question": row[0],
                    "choices": [
                        {"text": row[1]},
                        {"text": row[2]},
                        {"text": row[3]},
                        {"text": row[4]}
                    ],
                    "selected": row[5],
                    "is_right": row[6],  # 'Yes' or 'No'
                    "correct_answer": row[7]
                }
                questions.append(q_dict)

    except Exception as e:
        print("Error:", e)
        questions = []

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

    return render(request, "Test/candidate_answersheet.html", {
        "questions": questions,
        "candidate_id": candidate_id1,
        "candidate_name": candidate_name
    })                                                           