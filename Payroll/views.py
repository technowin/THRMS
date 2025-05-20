# views.py
from calendar import day_name
from decimal import Decimal
import json
# from tkinter import font
import math
import re
import os
import traceback
from colorama import Cursor
from django.conf import settings
from django.http import FileResponse, JsonResponse
from rest_framework.response import Response
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from pyparsing import str_type
from Account import models
from Account.models import CustomUser
from Masters.models import CityMaster, SlotDetails, StateMaster, UserSlotDetails, company_master, parameter_master, sc_employee_master, site_master
from Masters.serializers import PaySlipSerializer, SalaryGeneratedSerializer
from Payroll.forms import ExcelUploadForm, RateCardMasterForm, SalaryElementMasterForm, SiteCardRelationForm
from Payroll.models import IncomeTaxCalculation, IncomeTaxMaster, PaySlip, PayoutDetails, RateCardSalaryElement, PayrollStatusMaster, daily_salary, employee_rate_card_details, income_tax_deduction, income_tax_parameter
from Payroll.models import payment_details as pay, rate_card_master, salary_element_master, salary_generated_log, site_card_relation, slot_attendance_details
from THRMS.encryption import enc, dec
from django.contrib.auth.decorators import login_required
from datetime import timedelta
from django.db.models import Sum
from rest_framework.views import APIView
import pandas as pd
from django.views.generic import ListView
from datetime import datetime
from django.db.models import Case, When
import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from django.db import connection
import requests
from django.http import JsonResponse
import Db
from datetime import date
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa 
import base64
from io import BytesIO
from PIL import Image
from django.shortcuts import render
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework_simplejwt.tokens import AccessToken
from django.core.files.base import ContentFile
from xhtml2pdf import pisa
from io import BytesIO
from django.core.files.base import ContentFile
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse

from Tax.models import ActMaster, SlabMaster, TaxCalculation
# from openpyxl.styles import Font
# Index (list all salary elements)

def parse_challan_period(challan_period):
    try:
        if "to" in challan_period:  # Quarterly or Half-Yearly
            start, end = challan_period.split(" to ")
            start_month, start_year = map(int, start.split("-"))
            end_month, end_year = map(int, end.split("-"))
            return datetime(start_year, start_month, 1), datetime(end_year, end_month, 1)
        else:  # Monthly
            month, year = map(int, challan_period.split("-"))
            start_date = datetime(year, month, 1)
            return start_date, start_date
    except ValueError as e:
        raise ValueError("Invalid challan_period format") from e
    
def get_employer_deduct(comp_id_in, act_id_in, user_id_in, emp_code_in, 
                        city_id_in, state_id_in, site_id_in, month_in, year_in):
    
    employer_deduct = 0  # Default amount if no valid record is found
    
    # Step 1: Check record existence in tax_calculation
    record = TaxCalculation.objects.filter(
        act_id=act_id_in,
        city=city_id_in,
        state_id=state_id_in,
        status = 0,
        employee_id=emp_code_in
    ).first()
    
    if not record:
        return employer_deduct
    else:
        challan_period = record.challan_period # No record exists, return 0
    
    # Step 2: Parse and validate challan_period
    try:
        start_date, end_date = parse_challan_period(challan_period)
        challan_date = datetime(month_in, month_in, 1)

        if start_date <= challan_date <= end_date:
            print("Challan period already includes the given month and year.")
            return employer_deduct
    except ValueError as e:
        print(f"Error parsing challan_period: {e}")
    
    # Step 3: Get employer deduct from the record
    employer_deduct = record.employer_deduct  
    
    record.status = 1
    record.save(update_fields=["status"])  # Update only the status field to optimize the query
    
    return employer_deduct



# PT calculation logic
def calculate_professional_tax(user_session, emp_code_in, act_id_in, re_year_in, re_month_in, state_id_in, city_id_in, site_id_in, gross_salary_in, emp_gender_in, comp_id_in):
    try:
        slab = SlabMaster.objects.filter(
                state_id=state_id_in,
                city=city_id_in,
                act_id=act_id_in
            ).first()

            # If no slab is found, fall back to checking only state_id
        if not slab:
                slab = SlabMaster.objects.filter(
                    state_id=state_id_in,
                    act_id=act_id_in
                ).first()

        if not slab :
            return 0

        if slab:

            tax_record = TaxCalculation.objects.filter(
            employee_id=emp_code_in,
            state_id=state_id_in,
            city=city_id_in,
            act=act_id_in
            ).first()
    
            if tax_record:
            # Check existing slab_freq and process based on its value
                slab_freq = tax_record.slab_freq
                if slab_freq == "Monthly":
                    tax_record = TaxCalculation.objects.filter(
                        employee_id=emp_code_in,
                        state_id=state_id_in,
                        city=city_id_in,
                        act=act_id_in,
                        year=re_year_in,
                        month=re_month_in
                    ).first()
                    if tax_record is not None:
                        if int(tax_record.year) == int(re_year_in) and int(tax_record.month) == int(re_month_in):
                            tax_record.gross_salary += gross_salary_in  # Update the gross_salary field
                            gross_salary_in = tax_record.gross_salary  # Assign the updated value to gross_salary (if needed)

                    else:
                        tax_record = TaxCalculation.objects.create(
                            employee_id=emp_code_in,
                            state=get_object_or_404(StateMaster, state_id=state_id_in),
                            act=get_object_or_404(ActMaster, act_id=act_id_in),
                            city=city_id_in,
                            company=get_object_or_404(company_master, company_id=comp_id_in),
                            updated_by=get_object_or_404(CustomUser, id=user_session),
                            created_by=get_object_or_404(CustomUser, id=user_session),
                            gross_salary=gross_salary_in,
                            employee_tax=0,
                            challan_month=re_month_in,
                            challan_year=re_year_in,
                            challan_period=f"{re_month_in}-{re_year_in}",
                            year=re_year_in,
                            month=re_month_in,
                            slab_freq=slab_freq,
                            tax_deducted=0,
                            site=get_object_or_404(site_master, site_id=site_id_in)
                        )
                elif slab_freq == "Quarterly":
                    challan_period = tax_record.challan_period
                    start_date, end_date = parse_challan_period(challan_period)

                    # Check if a record exists for the given challan period
                    tax_record = TaxCalculation.objects.filter(
                        employee_id=emp_code_in,
                        state_id=state_id_in,
                        city=city_id_in,
                        act=act_id_in,
                        year__gte=start_date.year,
                        year__lte=end_date.year,
                        challan_period=f"{start_date.month}-{start_date.year} to {end_date.month}-{end_date.year}"
                    ).first()

                    if tax_record  and start_date <= date(re_year_in, re_month_in, 1) <= end_date:
                        tax_record.gross_salary += gross_salary_in
                        gross_salary_in = tax_record.gross_salary
                        tax_record.save()
                    else:
                        tax_record = TaxCalculation.objects.create(
                            employee_id=emp_code_in,
                            state=get_object_or_404(StateMaster, state_id=state_id_in),
                            act=get_object_or_404(ActMaster, act_id=act_id_in),
                            city=city_id_in,
                            company=get_object_or_404(company_master, company_id=comp_id_in),
                            updated_by=get_object_or_404(CustomUser, id=user_session),
                            created_by=get_object_or_404(CustomUser, id=user_session),
                            gross_salary=gross_salary_in,
                            employee_tax=0,
                            challan_period=challan_period,
                            year=re_year_in,
                            month=re_month_in,
                            slab_freq=slab_freq,
                            tax_deducted=0,
                            site=get_object_or_404(site_master, site_id=site_id_in)
                        )

                elif slab_freq == "Half Yearly":
                    challan_period = tax_record.challan_period
                    start_date, end_date = parse_challan_period(challan_period)

                    # Check if a record exists for the given challan period
                    tax_record = TaxCalculation.objects.filter(
                        employee_id=emp_code_in,
                        state_id=state_id_in,
                        act=act_id_in,
                        city=city_id_in,
                        year__gte=start_date.year,
                        year__lte=end_date.year,
                        challan_period=f"{start_date.month}-{start_date.year} to {end_date.month}-{end_date.year}"
                    ).first()

                    if tax_record and start_date <= date(re_year_in, re_month_in, 1) <= end_date:
                        tax_record.gross_salary += gross_salary_in
                        gross_salary_in = tax_record.gross_salary
                        tax_record.save()
                    else:
                        tax_record = TaxCalculation.objects.create(
                            employee_id=emp_code_in,
                            state=get_object_or_404(StateMaster, state_id=state_id_in),
                            act=get_object_or_404(ActMaster, act_id=act_id_in),
                            city=city_id_in,
                            company=get_object_or_404(company_master, company_id=comp_id_in),
                            updated_by=get_object_or_404(CustomUser, id=user_session),
                            created_by=get_object_or_404(CustomUser, id=user_session),
                            gross_salary=gross_salary_in,
                            employee_tax=0,
                            challan_period=challan_period,
                            year=re_year_in,
                            month=re_month_in,
                            slab_freq=slab_freq,
                            tax_deducted=0,
                            site=get_object_or_404(site_master, site_id=site_id_in)
                        )

                elif slab_freq == "Yearly":
                    challan_period = tax_record.challan_period
                    start_date, end_date = parse_challan_period(challan_period)

                    # Check if a record exists for the given challan period
                    tax_record = TaxCalculation.objects.filter(
                        employee_id=emp_code_in,
                        state_id=state_id_in,
                        act=act_id_in,
                        city=city_id_in,
                        year__gte=start_date.year,
                        year__lte=end_date.year,
                        challan_period=f"{start_date.month}-{start_date.year} to {end_date.month}-{end_date.year}"
                    ).first()

                    if tax_record and start_date <= date(re_year_in, re_month_in, 1) <= end_date:
                            tax_record.gross_salary += gross_salary_in
                            gross_salary_in = tax_record.gross_salary
                            tax_record.save()
                else:
                    tax_record = TaxCalculation.objects.create(
                        employee_id=emp_code_in,
                        state=get_object_or_404(StateMaster, state_id=state_id_in),
                        act=get_object_or_404(ActMaster, act_id=act_id_in),
                        city=city_id_in,
                        company=get_object_or_404(company_master, company_id=comp_id_in),
                        updated_by=get_object_or_404(CustomUser, id=user_session),
                        created_by=get_object_or_404(CustomUser, id=user_session),
                        gross_salary=gross_salary_in,
                        employee_tax=0,
                        challan_period=challan_period,
                        year=re_year_in,
                        month=re_month_in,
                        slab_freq=slab_freq,
                        tax_deducted=0,
                        site=get_object_or_404(site_master, site_id=site_id_in)
                    )
            # Additional logic for Quarterly, Half-Yearly, and Yearly can be added here
            else:
                # Create new record if not found
                tax_record = TaxCalculation.objects.create(
                employee_id=emp_code_in,
                state=get_object_or_404(StateMaster, state_id=state_id_in),
                act=get_object_or_404(ActMaster, act_id=act_id_in),
                city=city_id_in,
                company=get_object_or_404(company_master, company_id=comp_id_in),
                updated_by=get_object_or_404(CustomUser, id=user_session),
                created_by=get_object_or_404(CustomUser, id=user_session),
                gross_salary=gross_salary_in,
                employee_tax=0,
                challan_period=f"{re_month_in}-{re_year_in}",
                challan_month=re_month_in,
                challan_year=re_year_in,
                year=re_year_in,
                month=re_month_in,
                tax_deducted=0,
                site=get_object_or_404(site_master, site_id=site_id_in)
            )
    
            # Call pt_calculation
            (
                comp_id_out, state_id_out, city_id_out, act_id_out, site_id_out,
                re_year_out, re_month_out, emp_code_out, gross_salary_out,
                pt_actual_out, salary_period_out, slab_freq_out
            ) = pt_calculation(
                user_session, tax_record.company.company_id, tax_record.year,
                tax_record.month, tax_record.act.act_id, emp_gender_in,
                tax_record.employee_id, gross_salary_in, tax_record.city,
                tax_record.state_id, tax_record.site.site_id
            )
    
            # Tax deduction logic
            current_tax_deducted = tax_record.tax_deducted
    
            if current_tax_deducted == 0:
                return_value = pt_actual_out
            elif current_tax_deducted == pt_actual_out:
                return_value = 0
            elif current_tax_deducted < pt_actual_out:
                return_value = pt_actual_out - current_tax_deducted
            else:
                print("Invalid tax data: current tax deducted exceeds actual tax.")
                return_value = None
    
            # Update tax record with new values
            tax_record.gross_salary = gross_salary_out
            tax_record.employee_tax = pt_actual_out
            tax_record.challan_period = salary_period_out
            tax_record.slab_freq = slab_freq_out
            tax_record.tax_deducted = pt_actual_out
            tax_record.save()
    
            return return_value

    except Exception as e:
        print(f"Error in PT calculation: {e}")
        return None
print



def get_total_income(emp_code_in, re_year_in, re_month_in):
    try:
        re_year_in = int(re_year_in)
        start_date = date(re_year_in, 4, 1)  # April 1st of given year
        end_date = date(re_year_in + 1, 3, 31)  # March 31st of next year

        # Filter data based on the financial year range
        total_income = (
            daily_salary.objects.filter(
                employee_id=emp_code_in,
                attendance_date__gte=start_date,
                attendance_date__lte=end_date,
                attendance_date__month=re_month_in,
                tax_parameter=1
            ).aggregate(total=Sum('amount'))['total']
        )

        return float(total_income) if total_income is not None else 0.0
    except Exception as e:
        print(f"Error in Total income calculation: {e}")
        return None
    # Return 0 if no data is found

def get_total_80C(emp_code_in, re_year_in, re_month_in):
    re_year_in = int(re_year_in)
    start_date = date(re_year_in, 4, 1)  # April 1st of given year
    end_date = date(re_year_in + 1, 3, 31) # Inclusive range

    total_80C = (
        daily_salary.objects.filter(
            employee_id=emp_code_in,
                attendance_date__gte=start_date,
                attendance_date__lte=end_date,
                attendance_date__month=re_month_in,
            tax_parameter=2
        ).aggregate(total=Sum('amount'))['total']
    )
    return float(total_80C) if total_80C is not None else 0.0

def get_pt_value(emp_code_in, re_year_in, re_month_in):
    re_year_in = int(re_year_in)
    start_date = date(re_year_in, 4, 1)  # April 1st of given year
    end_date = date(re_year_in + 1, 3, 31) # Inclusive range

    total_pt = (
        daily_salary.objects.filter(
                employee_id=emp_code_in,
                attendance_date__gte=start_date,
                attendance_date__lte=end_date,
                attendance_date__month=re_month_in,
                tax_parameter=3
        ).aggregate(total=Sum('amount'))['total']
    )

    return float(total_pt) if total_pt is not None else 0.0


def calculate_lwf_tax(user_session, emp_code_in, act_id_in, re_year_in, re_month_in, state_id_in, city_id_in, site_id_in, gross_salary_in, emp_gender_in, comp_id_in):
    try:
        # Check if a record exists in TaxCalculation
        tax_record = TaxCalculation.objects.filter(
            employee_id=emp_code_in,
            state_id=state_id_in,
            city=city_id_in,
            act=act_id_in
        ).first()

        if tax_record:
        # Parse challan_period to check if re_month_in and re_year_in are within the range
            try:
                start_date, end_date = parse_challan_period(tax_record.challan_period)
                challan_date = datetime(re_year_in, re_month_in, 1)

                if start_date <= challan_date <= end_date:
                    print("Challan period already includes the given month and year.")
                    return 0
            except ValueError as e:
                print(f"Error parsing challan_period: {e}")
                return None

        if not tax_record:
            # If record doesn't exist, check in SlabMaster
            # First, try to get the slab with both state_id and city_id
            slab = SlabMaster.objects.filter(
                state_id=state_id_in,
                city_id=city_id_in,
                act_id=act_id_in
            ).first()

            # If no slab is found, fall back to checking only state_id
            if not slab:
                slab = SlabMaster.objects.filter(
                    state_id=state_id_in,
                    act_id=act_id_in
                ).first()

            # If still no slab is found, handle the absence of a valid slab
            if not slab:
                print("No slab found for the given criteria.")
                employee_deduct = 0
                employer_deduct = 0
                return employee_deduct,employer_deduct


            # Validate if month_in exists in slab_months_challan
            slab_months_list = list(map(int, slab.slab_months.split(',')))

            # Check if re_month_in exists in slab_months_list
            if re_month_in not in slab_months_list:
                print(f"Month {re_month_in} is not valid in slab_months.")
                employee_deduct = 0
                employer_deduct = 0
                return employee_deduct
            
            slab_id = slab.slab_id
            period = slab.period
            slab_freq = slab.slab_freq
            slab_months = slab.slab_months

            # Create a new record in TaxCalculation
            tax_record = TaxCalculation.objects.create(
                employee_id=emp_code_in,
                state=get_object_or_404(StateMaster, state_id=state_id_in),
                act=get_object_or_404(ActMaster, act_id=act_id_in),
                city=city_id_in,
                company=get_object_or_404(company_master, company_id=comp_id_in),
                updated_by=get_object_or_404(CustomUser, id=user_session),
                created_by=get_object_or_404(CustomUser, id=user_session),
                gross_salary=gross_salary_in,
                employee_tax=0,
                employer_deduct=0,
                challan_month=re_month_in,
                challan_year=re_year_in,
                challan_period=f"{re_month_in}-{re_year_in}",
                year=re_year_in,
                month=re_month_in,
                slab_freq=slab_freq,
                tax_deducted=0,
                site=get_object_or_404(site_master, site_id=site_id_in),
            )

        # Call lwf_calculation with appropriate parameters
        (
            comp_id_out, state_id_out, city_id_out, act_id_out, site_id_out,
            re_year_out, re_month_out, emp_code_out, gross_salary_out,
            setEmployee, setEmployer, challan_period
        ) = lwf_calculation(
            user_session,tax_record.company.company_id, tax_record.year,
            tax_record.month, tax_record.act.act_id, emp_gender_in,
            tax_record.employee_id, gross_salary_in, tax_record.city,
            tax_record.state_id, tax_record.site.site_id,slab_freq,period,slab_id,slab_months
        )

        # Update tax record with returned values
        tax_record.gross_salary = gross_salary_out
        tax_record.employee_tax = setEmployee

        tax_record.employer_deduct = setEmployer  # Example, update based on logic
        tax_record.challan_period = challan_period
        tax_record.tax_deducted = setEmployer
        tax_record.status = 0
        tax_record.save()

        return setEmployee

    except Exception as e:
        print(f"Error in LWF calculation: {e}")
        return None
    



def calculate_income_tax(user_session, emp_code_in, re_year_in, re_month_in,gross_earning):
    try:
        total_income = get_total_income(emp_code_in, re_year_in, re_month_in)
        total_80c = get_total_80C(emp_code_in, re_year_in, re_month_in)
        pt_value = get_pt_value(emp_code_in, re_year_in, re_month_in)

        financial_year_obj = parameter_master.objects.filter(parameter_name='financial year', is_active=1).first()
        if not financial_year_obj:
            return {'error': 'Financial year not found'}
        
        financial_year = financial_year_obj.parameter_value  
        input_period = f"4-{re_year_in} to 3-{re_year_in + 1}"

        standard_deduction = Decimal(parameter_master.objects.filter(parameter_name='Standard Deduction').first().parameter_value or '0')
        income_tax_standard = Decimal(parameter_master.objects.filter(parameter_name='Income Tax Standard').first().parameter_value or '0')
        edu_cess = Decimal(parameter_master.objects.filter(parameter_name='Edu Cess').first().parameter_value or '0')

        

        tax_record, created = IncomeTaxCalculation.objects.get_or_create(
            employee_id=emp_code_in,
            period=input_period,
            defaults={
                'financial_year': financial_year,
                'total_gross':gross_earning,
                'taxable_income': Decimal(total_income),
                'total_80c': Decimal(total_80c),
                'total_pt': Decimal(pt_value),
                'standard_deduction': standard_deduction,
                'created_by': get_object_or_404(CustomUser, id=user_session),
            }
        )

        if not created:
            old_total_gross = tax_record.total_gross or Decimal(0)
    
            tax_record.total_gross = old_total_gross + Decimal(gross_earning)

            tax_record.taxable_income += Decimal(total_income)
            tax_record.total_80c += Decimal(total_80c)
            tax_record.total_pt += Decimal(pt_value)

            tax_record.updated_by = get_object_or_404(CustomUser, id=user_session)
            tax_record.save()


        print(f"Taxable Income: {tax_record.total_gross}")


        taxable_income = tax_record.total_gross - standard_deduction
        tax_record.total_income = taxable_income
        tax_record.tax_paid = 0
        tax_record.save()
        taxable_income = tax_record.total_income
        if taxable_income < 0:
            tax_record.tax_amount = Decimal('0')
            tax_record.total_tax_amount = Decimal('0')
            tax_record.save()
            return Decimal('0')

        taxable_income -= tax_record.total_pt
        if taxable_income < 0:
            tax_record.tax_amount = Decimal('0')
            tax_record.total_tax_amount = Decimal('0')
            tax_record.save()
            return Decimal('0')

        taxable_income -= tax_record.total_80c
        if taxable_income < 0:
            tax_record.tax_amount = Decimal('0')
            tax_record.total_tax_amount = Decimal('0')
            tax_record.save()
            return Decimal('0')

        if taxable_income <= income_tax_standard:
            tax_record.tax_amount = Decimal('0')
            tax_record.rebate_under_87a = Decimal('0')
            tax_record.save()
            return Decimal('0')

        # Tax Calculation
        remaining_income = taxable_income
        total_tax = Decimal('0')
        tax_slabs = IncomeTaxMaster.objects.filter(financial_year=financial_year).order_by('id')
        cumulative_slab_to = Decimal('0')

        for index, slab in enumerate(tax_slabs):
            slab_from = Decimal(slab.tax_slab_from)
            slab_to = Decimal(slab.tax_slab_to) if slab.tax_slab_to else Decimal('Infinity')
            tax_rate = Decimal(slab.tax_rate) / Decimal('100')
            edu_cess_percent = Decimal(edu_cess) / Decimal('100')

            if index == 0:
                slab_income = min(remaining_income, slab_to)
            else:
                adjusted_slab_limit = slab_to - cumulative_slab_to
                slab_income = min(remaining_income, adjusted_slab_limit)

            slab_income = max(Decimal('0'), slab_income)
            slab_tax = slab_income * tax_rate
            
            total_tax += slab_tax 

            # Update cumulative slab limit
            cumulative_slab_to = slab_to
            remaining_income -= slab_income

            if remaining_income <= 0:
                break

        edu_cess_amount = total_tax * edu_cess_percent
        tax = total_tax + edu_cess_amount
        tax_record.last_tax_paid = date.today()

        tax_record.save()

        # Calculate the income tax
        tax_paid = tax_record.tax_paid
        income_tax = tax - tax_paid  # This gives the remaining tax after deducting tax paid

        tax_record.tax_paid = income_tax

        # Update the tax_record with remaining taxable income
        tax_record.remaining_taxable_income = income_tax

        # Save the updated record
        tax_record.save()


        return float(income_tax)

    except Exception as e:
        return {'error': str(e)}

    





@login_required
def index(request):
    salary_elements = salary_element_master.objects.all()
    for record in salary_elements:
            # print(record.id)
        record.pk = enc(str(record.item_id))
    return render(request, 'Payroll/SalaryElement/index.html', {'salary_elements': salary_elements})

# Create a new salary element
@login_required
def create(request):
    if request.method == 'POST':
        form = SalaryElementMasterForm(request.POST)
        if form.is_valid():
            salary_element = form.save(commit=False)  # Don't save to the database yet
            salary_element.created_by = request.user  # Set created_by to the current user
            salary_element.updated_by = request.user  # Also set updated_by initially
            salary_element.save()  # Now save to the database
            messages.success(request, "Salary Element created successfully!")
            return redirect('salary_element_index')
        else:
            messages.error(request, "Error creating Salary Element.")
    else:
        form = SalaryElementMasterForm()
    
    return render(request, 'Payroll/SalaryElement/create.html', {'form': form})

# Edit an existing salary element
@login_required
def edit(request, pk):
    pk = dec(pk)
    salary_element = get_object_or_404(salary_element_master, pk=pk)
    
    if request.method == 'POST':
        form = SalaryElementMasterForm(request.POST, instance=salary_element)
        if form.is_valid():
            salary_element = form.save(commit=False)  # Don't save to the database yet
            salary_element.updated_by = request.user  # Also set updated_by initially
            salary_element.save()  # Now save to the database
            messages.success(request, "Salary Element updated successfully!")
            return redirect('salary_element_index')
        else:
            messages.error(request, "Error updating Salary Element.")
    else:
        form = SalaryElementMasterForm(instance=salary_element)
    
    return render(request, 'Payroll/SalaryElement/edit.html', {'form': form, 'salary_element': salary_element})

# View salary element details
@login_required
def view(request, pk):
    pk = dec(pk)
    salary_element = get_object_or_404(salary_element_master, pk=pk)
    return render(request, 'Payroll/SalaryElement/view.html', {'salary_element': salary_element})


@login_required
def rate_card_index(request):
    rate_cards = rate_card_master.objects.all()
    for record in rate_cards:
        record.pk = enc(str(record.card_id))
    return render(request, 'Payroll/RateCard/index.html', {'rate_cards': rate_cards})

@login_required
def rate_card_create(request):
    if request.method == "POST":
        form = RateCardMasterForm(request.POST)
        if form.is_valid():
            # Save the rate card instance first
            rate_card = form.save(commit=False)
            rate_card.created_by = request.user
            rate_card.updated_by = request.user
            rate_card.is_active = True
            rate_card.save()
            
            # Save the many-to-many relation with additional fields in the through model
            selected_items = request.POST.getlist('item_ids')  # Get selected item_ids
            
            for item_id in selected_items:
                item = salary_element_master.objects.get(pk=item_id)
                four_hour_amount = request.POST.get(f'four_hour_amount_{item_id}', 0)
                nine_hour_amount = request.POST.get(f'nine_hour_amount_{item_id}', 0)
                tax_param = request.POST.get(f'tax_parameter_{item_id}', None)
                
                if not tax_param or tax_param == '0':
                    tax_object = None
                else:
                    tax_object = get_object_or_404(income_tax_parameter, id=tax_param)
                RateCardSalaryElement.objects.create(
                    rate_card=rate_card,
                    salary_element=item,
                    item_name=item.item_name,
                    pay_type=item.pay_type,
                    classification=item.classification,
                    four_hour_amount=four_hour_amount,
                    nine_hour_amount=nine_hour_amount,
                    tax_parameter=tax_object
                )

            messages.success(request, 'Rate Card created successfully!')
            return redirect('rate_card_index')
        else:
            messages.error(request, 'Error creating Rate Card.')
    else:
        tax_parameter = income_tax_parameter.objects.all()
        form = RateCardMasterForm()
    
    return render(request, 'Payroll/RateCard/create.html', {'form': form,'tax_parameter':tax_parameter})



 
@login_required
def rate_card_edit(request, card_id):
    card_id = dec(card_id)
    rate_card = get_object_or_404(rate_card_master, pk=card_id)

    if request.method == "POST":
        form = RateCardMasterForm(request.POST, instance=rate_card)
        if form.is_valid():
            rate_card = form.save(commit=False)
            rate_card.updated_by = request.user
            rate_card.save()

            RateCardSalaryElement.objects.filter(rate_card=rate_card).delete()

            selected_items = request.POST.getlist('item_ids')

            for item_id in selected_items:
                item = salary_element_master.objects.get(pk=item_id)
                four_hour_amount = request.POST.get(f'four_hour_amount_{item_id}', 0)
                nine_hour_amount = request.POST.get(f'nine_hour_amount_{item_id}', 0)
                tax_param = request.POST.get(f'tax_parameter_{item_id}', 0)


                RateCardSalaryElement.objects.create(
                    rate_card=rate_card,
                    salary_element=item,
                    item_name=item.item_name,
                    pay_type=item.pay_type,
                    classification=item.classification,
                    four_hour_amount=four_hour_amount,
                    nine_hour_amount=nine_hour_amount,
                    tax_parameter=get_object_or_404(tax_parameter, id = tax_param)

                )

            messages.success(request, 'Rate Card updated successfully!')
            return redirect('rate_card_index')
        else:
            messages.error(request, 'Error updating Rate Card.')
    else:
        tax_parameter = income_tax_parameter.objects.all()
        form = RateCardMasterForm(instance=rate_card)

    selected_item_ids = rate_card.item_ids.values_list('item_id', flat=True)

    # Prepare pre-filled data for each item
    existing_relations = RateCardSalaryElement.objects.filter(rate_card=rate_card)

    prefilled_data = {}
    for relation in existing_relations:
        prefilled_data[relation.salary_element_id] = {
            'four_hour_amount': relation.four_hour_amount,
            'nine_hour_amount': relation.nine_hour_amount,
            'tax_parameter':relation.tax_parameter
        }

    return render(request, 'Payroll/RateCard/edit.html', {
        'form': form,
        'tax_parameter':tax_parameter,
        'rate_card': rate_card,
        'selected_item_ids': list(selected_item_ids),
        'prefilled_data': prefilled_data,  # Pass the prefilled amounts
    })

 
@login_required
def rate_card_view(request, card_id):
    card_id = dec(card_id)
    rate_card = get_object_or_404(rate_card_master, pk=card_id)
    salary_elements = RateCardSalaryElement.objects.filter(rate_card=rate_card)
    
    return render(request, 'Payroll/RateCard/view.html', {
        'rate_card': rate_card,
        'salary_elements': salary_elements,
    })


@login_required
def site_card_relation_index(request):
    site_card_relations = site_card_relation.objects.all()
    for record in site_card_relations:
        record.pk = enc(str(record.relation_id))
    return render(request, 'Payroll/SiteCardRelation/index.html', {'site_card_relations': site_card_relations})
@login_required
def site_card_relation_create(request):
    user = request.session.get('user_id', '')  # Retrieve the logged-in user's ID

    if request.method == 'POST':
        form = SiteCardRelationForm(request.POST, user_id=user)  # Pass user_id to the form
        if form.is_valid():
            form.save()
            messages.success(request, 'Site Card Relation created successfully!')
            return redirect('site_card_relation_index')
    else:
        form = SiteCardRelationForm(user_id=user)  # Pass user_id to the form

    return render(request, 'Payroll/SiteCardRelation/create.html', {'form': form})

@login_required
def site_card_relation_edit(request, pk):
    pk = dec(pk)
    site_card_relation_instance = get_object_or_404(site_card_relation, pk=pk)
    if request.method == 'POST':
        form = SiteCardRelationForm(request.POST, instance=site_card_relation_instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Site Card Relation updated successfully!')
            return redirect('site_card_relation_index')
    else:
        form = SiteCardRelationForm(instance=site_card_relation_instance)

    return render(request, 'Payroll/SiteCardRelation/edit.html', {'form': form})
@login_required
def site_card_relation_view(request, pk):
    pk = dec(pk)
    site_card_relation_instance = get_object_or_404(site_card_relation, pk=pk)
    return render(request, 'Payroll/SiteCardRelation/view.html', {'site_card_relation': site_card_relation_instance})
 
@login_required
def employee_rate_card_index(request):
    rate_cards = employee_rate_card_details.objects.all()
    return render(request, 'Payroll/EmployeeRateCardDetails/index.html', {'rate_cards': rate_cards})
 
@login_required
def employee_rate_card_create(request):
    try:
        # Fetch distinct site_ids and their names
        sites = site_card_relation.objects.all().values('site_id', 'site_id__site_name').distinct()
        
        # Fetch employees not in any rate card
        employees_not_in_rate_card = sc_employee_master.objects.exclude(employee_id__in=employee_rate_card_details.objects.values('employee_id'))

        if request.method == 'POST':
            selected_site_id = request.POST.get('site_id')
            selected_card_id = request.POST.get('card_id')
            selected_employees = request.POST.getlist('selected_employees')  # List of selected employee IDs
            
            # Get the selected rate card
            rate_card = rate_card_master.objects.get(card_id=selected_card_id)

            # Get all RateCardSalaryElement entries for the selected rate card
            rate_card_salary_elements = RateCardSalaryElement.objects.filter(rate_card=rate_card)

            # Loop through each selected employee
            for employee_id in selected_employees:
                # Loop through each salary element related to the rate card
                for rate_card_salary_element in rate_card_salary_elements:
                    employee_rate_card_details.objects.create(
                        employee_id=employee_id,                        # Employee ID from the selected list
                        card_id=rate_card,                              # Foreign key to rate_card_master
                        item_id=rate_card_salary_element.salary_element,  # Foreign key to salary_element_master
                        item_name=rate_card_salary_element.item_name,     # Item name from RateCardSalaryElement
                        pay_type=rate_card_salary_element.pay_type,       # Pay type from RateCardSalaryElement
                        classification=rate_card_salary_element.classification,  # Classification from RateCardSalaryElement
                        four_hour_amount=rate_card_salary_element.four_hour_amount,  # Four hour amount from RateCardSalaryElement
                        nine_hour_amount=rate_card_salary_element.nine_hour_amount,    # Nine hour amount from RateCardSalaryElement
                        created_by=request.user,                        # Assuming current user is creating the record
                        updated_by=request.user 
                    )
            return redirect('employee_rate_card_index')

        return render(request, 'Payroll/EmployeeRateCardDetails/create.html', {
            'sites': sites,
            'employees': employees_not_in_rate_card
        })
    except Exception as e:
        print(str(e))
        return redirect('employee_rate_card_index')
@login_required
def get_rate_cards(request, site_id):
    try:
        rate_cards = site_card_relation.objects.filter(site_id=site_id).select_related('card_id').values(
            'card_id', 'card_id__card_name', 'card_id__is_active', 'card_id__created_at'
        )
        
        cards = [
            {
                'card_id': rc['card_id'],
                'card_name': rc['card_id__card_name'],
                'is_active': rc['card_id__is_active'],
                'created_at': rc['card_id__created_at'],
            } for rc in rate_cards
        ]
        return JsonResponse({'cards': cards})
    except Exception as e:
        print(str(e))
        return JsonResponse({'a':"a"})
@login_required
def employee_rate_card_edit(request, id):
    rate_card_detail = get_object_or_404(employee_rate_card_details, id=id)
    sites = site_card_relation.objects.all().values('site_id', 'site_id__site_name').distinct()
    employees = sc_employee_master.objects.all()

    if request.method == 'POST':
        rate_card_detail.employee_id = request.POST.get('employee_id')
        rate_card_detail.card_id_id = request.POST.get('card_id')  # Store card_id as ForeignKey ID
        rate_card_detail.item_id_id = request.POST.get('item_id')  # Store item_id as ForeignKey ID
        rate_card_detail.four_hour_amount = request.POST.get('four_hour_amount')
        rate_card_detail.nine_hour_amount = request.POST.get('nine_hour_amount')
        rate_card_detail.save()

        return redirect('employee_rate_card_index')

    return render(request, 'Payroll/EmployeeRateCardDetails/edit.html', {
        'rate_card_detail': rate_card_detail,
        'sites': sites,
        'employees': employees
    })
@login_required
def employee_rate_card_view(request, id):
    rate_card_detail = get_object_or_404(employee_rate_card_details, id=id)
    return render(request, 'Payroll/EmployeeRateCardDetails/view.html', {
        'rate_card_detail': rate_card_detail
    })



 
@login_required
def attendance_index(request):
    try:
        user = request.session.get('user_id', '')
        type = request.GET.get('type', '')
        attendance_records = slot_attendance_details.objects.filter(
            # company_id__in=user_role_map.objects.filter(
                user_id=user
            ).values('company_id')
        

        
        # Encrypt the ID for each record
        for record in attendance_records:
            # print(record.id)
            record.encrypted_id = enc(str(record.id))
        
        return render(request, 'Payroll/Attendance/index.html', {
            'attendance_records': attendance_records,
            'type': type
        })
    except Exception as e:
        print(f"Error: {e}") 
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        messages.error(request, 'Oops...! Something went wrong!')
    
    # type = request.GET.get('type', '')
    # attendance_records = slot_attendance_details.objects.all()
    # return render(request, 'Payroll/Attendance/index.html', {'attendance_records': attendance_records,'type':type})
 
# @login_required
# def upload_attendance(request):
#     if request.method == 'POST':
#         excel_form = ExcelUploadForm(request.POST, request.FILES)
#         if excel_form.is_valid():
#             excel_file = request.FILES['excel_file']
#             data = pd.read_excel(excel_file)
#             comp = company_master.objects.get(company_id=request.POST['company_id'])
#             site = site_master.objects.get(site_id=request.POST['site_id'])
#             slot = SlotDetails.objects.get(slot_id=request.POST['slot_id'])
#             for index, row in data.iterrows():
#                 attendance = slot_attendance_details(
#                     company_id= comp ,
#                     site_id= site,
#                     slot_id= slot,
#                     attendance_date=row['Attendance Date'],
#                     employee_id=row['Employee Id'],
#                     attendance_in=row['Attendance In'],
#                     attendance_out=row['Attendance Out'],
#                 )
#                 attendance.save()
#             messages.success(request, 'Attendance records uploaded successfully.')
#             return redirect('attendance_index')

#     else:
#         excel_form = ExcelUploadForm()

#     return render(request, 'Payroll/Attendance/create.html', {
#         'excel_form': excel_form,
#         'companies': company_master.objects.all(),
#     })

def format_time(time_str):
    """Ensure time format is HH:MM:SS by appending :00 if necessary."""
    time_str = str(time_str).strip()
    if re.match(r'^\d{2}:\d{2}$', time_str):  # Matches HH:MM format
        return time_str + ':00'
    return time_str


@login_required
def upload_attendance(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    user = request.session.get('user_id', '')

    try:
        if request.method == 'POST':
            excel_form = ExcelUploadForm(request.POST, request.FILES)
            if excel_form.is_valid():
                excel_file = request.FILES['excel_file']
                file_name = excel_file.name
                
                # Skip metadata rows (first 2 rows), read data from the third row onward
                data = pd.read_excel(excel_file)

                # Retrieve related objects
                comp = get_object_or_404(company_master, company_id=request.POST['company_id'])
                site = get_object_or_404(site_master, site_id=request.POST['site_id'])
                slot = get_object_or_404(SlotDetails, slot_id=request.POST['slot_id'])
                upload_for = 'Attendance Upload'
                
                # Insert checksum record
                cursor.callproc('stp_insert_checksum', (
                    upload_for, comp.company_id, str(datetime.now().month), str(datetime.now().year), file_name
                ))
                checksum_id = list(cursor.stored_results())[0].fetchall()[0][0]

                error_count = 0
                success_count = 0
                total_columns = len(data.columns)
                total_rows = len(data)
                
                for index, row in data.iterrows():
                    row_error_found = False

                    date_str = str(row.get('Attendance Date', '')).strip()

                    # If the date is provided, try parsing it
                    if date_str:
                        attendance_date = None

                        # Check if the date is already a datetime object (in case Excel passed a date object)
                        if isinstance(row.get('Attendance Date'), datetime):
                            attendance_date = row['Attendance Date'].date()
                        else:
                            # Try parsing 'dd-mm-yyyy' format
                            if '-' in date_str and len(date_str.split('-')[0]) <= 2:  # Likely 'dd-mm-yyyy'
                                try:
                                    attendance_date = datetime.strptime(date_str, '%d-%m-%Y').date()  # Extract only date
                                except ValueError:
                                    pass

                            # Try parsing 'yyyy-mm-dd' format
                            if not attendance_date and '-' in date_str and len(date_str.split('-')[0]) == 4:  # Likely 'yyyy-mm-dd'
                                try:
                                    attendance_date = datetime.strptime(date_str, '%Y-%m-%d').date()  # Extract only date
                                except ValueError:
                                    pass

                        # If parsing succeeds, continue processing with attendance_date
                        if attendance_date:
                            pass
                        else:
                            error_message = f"Invalid date format for Employee ID {row['Employee Id']} on {date_str}"
                            cursor.callproc('stp_insert_attendance_error_log', [
                                upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user
                            ])
                            error_count += 1
                            row_error_found = True
                            continue
                    else:
                        error_message = f"Attendance Date is missing for Employee ID {row['Employee Id']}"
                        cursor.callproc('stp_insert_attendance_error_log', [
                            upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user
                        ])
                        error_count += 1
                        row_error_found = True
                        continue


                    # Validate Employee ID (check if it is blank, null, or ' ')
                    if not row['Employee Id'] or str(row['Employee Id']).strip() == '':
                        error_message = f"Employee Id is missing or invalid at row number {row + 1}"
                        cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                        error_count += 1
                        row_error_found = True
                        continue

                    # Check if Employee ID exists in sc_employee_master
                    try:
                        employee = sc_employee_master.objects.get(employee_id=row['Employee Id'])
                    except sc_employee_master.DoesNotExist:
                        error_message = f"No employee with ID {row['Employee Id']}"
                        cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                        error_count += 1
                        row_error_found = True
                        continue

                    
                    try:
                        employee_id = row.get('Employee Id')
                        employee = sc_employee_master.objects.get(employee_id=employee_id, company_id=comp)
                    except sc_employee_master.DoesNotExist:
                        error_message = f"Employee with ID {employee_id} does not belong to company name {comp.company_name}."
                        cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, employee_id, user])
                        error_count += 1
                        row_error_found = True
                        return error_count, row_error_found

                    # Validate Attendance In

                    if not row['Attendance In'] or str(row['Attendance In']).strip() == '':
                        error_message = f"Attendance In is missing for Employee ID {row['Employee Id']}"
                        cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                        error_count += 1
                        row_error_found = True
                        continue
                    else:
                        row['Attendance In'] = format_time(row['Attendance In'])  # Format time before inserting

                    if not row['Attendance Out'] or str(row['Attendance Out']).strip() == '':
                        error_message = f"Attendance Out is missing for Employee ID {row['Employee Id']}"
                        cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                        error_count += 1
                        row_error_found = True
                        continue
                    else:
                        row['Attendance Out'] = format_time(row['Attendance Out'])
                    # if not row['Attendance In'] or str(row['Attendance In']).strip() == '':
                    #     error_message = f"Attendance In is missing for Employee ID {row['Employee Id']}"
                    #     cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                    #     error_count += 1
                    #     row_error_found = True
                    #     continue

                    # if not row['Attendance Out'] or str(row['Attendance Out']).strip() == '':
                    #     error_message = f"Attendance out is missing for Employee ID {row['Employee Id']}"
                    #     cursor.callproc('stp_insert_attendance_error_log', [upload_for, comp.company_id, file_name, error_message, site.site_id, checksum_id, row['Employee Id'], user])
                    #     error_count += 1
                    #     row_error_found = True
                    #     continue


                    attendance = slot_attendance_details(
                        company_id=comp,
                        site_id=site,
                        slot_id=slot,
                        attendance_date=attendance_date,
                        employee_id=row['Employee Id'],
                        attendance_in=row['Attendance In'],
                        attendance_out=row['Attendance Out'],
                        status_id=get_object_or_404(PayrollStatusMaster, status_id=1).status_id,
                        created_by = get_object_or_404(CustomUser, id = user)
                    )
                    
                    attendance.save()


                    # Update the status in the UserSlotDetails table
                    user_slot = UserSlotDetails.objects.filter(slot_id=slot, employee_id=row['Employee Id']).first()

                    if user_slot:
                        # Update the status if the record exists
                        user_slot.status_id = get_object_or_404(PayrollStatusMaster, status_id=1).status_id
                        user_slot.save()
    

                    # user_slot_details = get_object_or_404(UserSlotDetails, slot_id = slot.slot_id, employee_id =) 
                    # user_slot_details.status = get_object_or_404(PayrollStatusMaster, status_id=1) 

                    result_value = "success"  # Example, you should fetch this from the actual procedure result

                    if result_value == "success":
                        success_count += 1
                        slot_details = get_object_or_404(SlotDetails, slot_id=slot.slot_id)
                        slot_details.status = get_object_or_404(PayrollStatusMaster, status_id=1)  # Assuming status_id is a ForeignKey
                        slot_details.save()
                    elif result_value.startswith("error"):
                        # Log the error message in your error log table
                        error_message = result_value  # The error message from the procedure
                        cursor.callproc('stp_insert_error_log', [
                            upload_for, comp.company_id, 'attendance_file', datetime.now().date(), error_message, None, row['Employee Id']
                        ])
                        error_count += 1 


                # Prepare checksum message
                checksum_msg = (
                    f"Total Columns Processed in each row: {total_columns}, Total Rows Processed: {total_rows}, "
                    f"Successful Entries: {success_count}"
                    f"{f', Errors: {error_count}' if error_count > 0 else ''}"
                )

                # Update checksum status
                cursor.callproc('stp_update_checksum_attendance', (
                    upload_for, comp.company_id,site.site_name, str(datetime.now().month), str(datetime.now().year),
                    file_name, checksum_msg, error_count, checksum_id
                ))

                if error_count == 0  and success_count > 0:
                    messages.success(request, "All data uploaded successfully!")
                else:
                    messages.warning(request,
                        f"The upload processed {total_columns} columns, {total_rows} rows, resulting in {success_count} successful entries, please check the error logs for error :{error_count}."
                    )

                return redirect('/attendance_index?type=index')

        else:
            excel_form = ExcelUploadForm()
            # company = company_master.objects.filter(
            #     company_id__in=user_role_map.objects.filter(user_id=user).values_list('company_id', flat=True)
            # )

        # Handling exceptions and cursor closing
    except Exception as e:
        print(f"Error: {e}") 
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])
        messages.error(request, 'Oops...! Something went wrong!')

    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()

    return render(request, 'Payroll/Attendance/create.html', {
        'type': type,
        'excel_form': excel_form,
        # 'companies': company,
    })




@login_required
def get_sites(request):
    company_id = request.GET.get('company_id')
    sites = site_master.objects.filter(company_id=company_id).values('site_id', 'site_name')
    return JsonResponse(list(sites), safe=False)
@login_required
def get_slots(request):
    company_id = request.GET.get('company_id')
    site_id = request.GET.get('site_id')
    slots = SlotDetails.objects.filter(site_id=site_id, company_id=company_id).values('slot_id', 'slot_name')
    return JsonResponse(list(slots), safe=False)



 
@login_required
def calculate_daily_salary(request,slot_id):
    user_session = request.session.get('user_id', '') 
    # Step 1: Fetch employees for the given slot_id from slot_employee_details

    # Step 2: Get the slot details
    slot = SlotDetails.objects.get(slot_id=slot_id)
    site_ins = slot.site_id  # This references the related SiteMaster o
    state_id = site_ins.state_id_id
    city_id = site_ins.city_id_id
    print(slot.night_shift)
    employees = UserSlotDetails.objects.filter(slot_id=slot_id)
    user_slot_details = UserSlotDetails.objects.filter(slot_id=slot_id)
    
    print(employees)
    generated_logs = salary_generated_log.objects.filter(
    slot_id=slot_id,
    slot_date=slot.shift_date
    ).values_list('employee_id', 'company_id')  # Get both employee_id and company_id as tuples
    print(generated_logs)
    # Exclude employees where the combination of employee_id and company_id is in the generated_logs
    filtered_employees = employees.exclude(
        Q(employee_id__in=[log[0] for log in generated_logs]) & 
        Q(company_id__in=[log[1] for log in generated_logs])
    )
    print(filtered_employees)
    for employee in filtered_employees:
        employee_id = employee.employee_id

           
        
        # Step 3: Check attendance for the employee in the given slot
        attendance = slot_attendance_details.objects.filter(
            site_id=slot.site_id,
            slot_id=slot_id,
            employee_id=employee_id,
            attendance_date = slot.shift_date,
            status = 2

        ).first()
        
        if attendance:
            # Step 4: Calculate the total working hours
            if attendance.attendance_in and attendance.attendance_out:
                if ':' in attendance.attendance_in and ':' in attendance.attendance_out:
                    try:
                        # Determine the correct format
                        time_format = "%H:%M:%S" if len(attendance.attendance_in.split(":")) == 3 else "%H:%M"
                        
                        # Convert to time objects
                        time_in = datetime.strptime(attendance.attendance_in, time_format).time()
                        time_out = datetime.strptime(attendance.attendance_out, time_format).time()

                        # Calculate working hours
                        time_in_seconds = time_in.hour * 3600 + time_in.minute * 60
                        time_out_seconds = time_out.hour * 3600 + time_out.minute * 60

                        # Handling case where time_out might be on the next day (crossing midnight)
                        if time_out_seconds < time_in_seconds:
                            time_out_seconds += 24 * 3600  # Add 24 hours in seconds to time_out

                        # Calculate the total working hours
                        working_hours = (time_out_seconds - time_in_seconds) / 3600

                        print(f"Total working hours: {working_hours}")

                    except ValueError as e:
                        print(f"Error parsing time: {e}")
                        # Step 5: Fetch the rate card from site_card_relation based on site and designation
                    if working_hours !=0:
                        site_card_relation_obj = site_card_relation.objects.filter(
                            site_id=slot.site_id,
                            designation_id=slot.designation_id
                        ).first()

                        if site_card_relation_obj:
                            # Get the rate card ID and related salary elements
                            card_id = site_card_relation_obj.card_id
                            salary_elements = RateCardSalaryElement.objects.filter(rate_card=card_id).annotate(
                            pay_type_order=Case(
                                When(pay_type='Earning', then=1),
                                When(pay_type='Other Earning', then=2),
                                When(pay_type='Total Earning', then=3),
                                When(pay_type='Deduction', then=4),
                                When(pay_type='Income Tax', then=5),
                                When(pay_type='Employer Contribution', then=6),
                                When(pay_type='Employer Total', then=7),
                                When(pay_type='Total', then=8),
                                default=8,  # For any other pay types that are not specified
                                output_field=models.IntegerField(),
                                )
                            ).order_by('pay_type_order', 'item_name')

                            # Step 6: Fetch the BASIC element before looping
                            basic_element = salary_elements.filter(item_name='BASIC').first()
                            basic_amount = 0
                            amount = 0

                            # Determine BASIC amount based on working hours
                            if basic_element:
                                if working_hours < 9:
                                    basic_amount = basic_element.four_hour_amount
                                else:
                                    basic_amount = basic_element.nine_hour_amount
                                
                            daily_salary.objects.filter(slot_id=slot,
                                employee_id = employee_id,
                                company_id = employee.company_id,
                                attendance_date=attendance.attendance_date).delete()
                            # Step 7: Process each salary element
                            for element in salary_elements:
                                # Step 8: Handle Percentage-based calculations
                                if element.classification == 'Percentage':
                                    if element.item_name == 'ESIC Employee':
                                        gross_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            element_name='Gross Earning',  # Add filter for element_name
                                            pay_type='Total Earning'  # Retain pay_type filter if necessary
                                        ).aggregate(
                                            total_gross_earning=Sum('amount')  # Use Sum to aggregate the 'amount' column
                                        )['total_gross_earning'] or 0
                                        # Calculate percentage based on BASIC
                                        if working_hours < 9:
                                            amount = (gross_earning * element.four_hour_amount) / 100
                                        else:
                                            amount = (gross_earning * element.nine_hour_amount) / 100
                                    elif element.item_name == 'ESIC Employer':
                                        gross_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            element_name='Gross Earning',  # Add filter for element_name
                                            pay_type='Total Earning'  # Retain pay_type filter if necessary
                                        ).aggregate(
                                            total_gross_earning=Sum('amount')  # Use Sum to aggregate the 'amount' column
                                        )['total_gross_earning'] or 0
                                        # Calculate percentage based on BASIC
                                        if working_hours < 9:
                                            amount = (gross_earning * element.four_hour_amount) / 100
                                        else:
                                            amount = (gross_earning * element.nine_hour_amount) / 100
                                            
                                        # Calculate percentage based on BASIC
                                    else:
                                        # For other percentage-based items, use the normal logic
                                        if working_hours < 9:
                                            amount = (basic_amount * element.four_hour_amount) / 100
                                        else:
                                            amount = (basic_amount * element.four_hour_amount) / 100
                                
                                elif element.classification == "Calculation":
                                    if element.item_name == 'LWF Employee':
                                        gross_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            element_name='Gross Earning',  # Add filter for element_name
                                            pay_type='Total Earning'  # Retain pay_type filter if necessary
                                        ).aggregate(
                                            total_gross_earning=Sum('amount')  # Use Sum to aggregate the 'amount' column
                                        )['total_gross_earning'] or 0
                                    
                                        shift_date = slot.shift_date
                                        re_year_in = shift_date.year
                                        re_month_in = shift_date.month

                                        employee_id = employee.employee_id
                                        try:
                                            gender = sc_employee_master.objects.get(employee_id=employee_id).gender
                                            print(f"Gender for employee {employee_id}: {gender}")
                                        except sc_employee_master.DoesNotExist:
                                            print(f"No employee found with ID {employee_id}")

                                        comp_id_in = slot.company.company_id
                                        act_id_in = 2
                                        user_id_in = user_session
                                        emp_code_in = employee_id
                                        gross_salary_in = gross_earning
                                        emp_gender_in = gender
                                        city_id_in = city_id
                                        state_id_in = state_id
                                        site_id_in = slot.site_id.site_id  


                                        setEmployee= calculate_lwf_tax(user_session, emp_code_in, act_id_in, re_year_in, re_month_in, state_id_in, city_id_in, site_id_in, gross_salary_in,emp_gender_in,comp_id_in)

                                        if working_hours < 9:
                                            amount = setEmployee
                                        else:
                                            amount = setEmployee
                                        # Calculate percentage based on BASIC
                                        
                                        # if working_hours < 9:
                                        #     amount = (basic_amount * element.four_hour_amount) / 100
                                        # else:
                                        #     amount = (basic_amount * element.nine_hour_amount) / 100
                                    elif element.item_name == 'Professional Tax':
                                        gross_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            element_name='Gross Earning',  # Add filter for element_name
                                            pay_type='Total Earning'  # Retain pay_type filter if necessary
                                        ).aggregate(
                                            total_gross_earning=Sum('amount')  # Use Sum to aggregate the 'amount' column
                                        )['total_gross_earning'] or 0
                                    
                                        shift_date = slot.shift_date
                                        re_year_in = shift_date.year
                                        re_month_in = shift_date.month

                                        employee_id = employee.employee_id
                                        try:
                                            gender = sc_employee_master.objects.get(employee_id=employee_id).gender
                                            print(f"Gender for employee {employee_id}: {gender}")
                                        except sc_employee_master.DoesNotExist:
                                            print(f"No employee found with ID {employee_id}")

                                        comp_id_in = slot.company.company_id
                                        act_id_in = 1
                                        user_id_in = user_session
                                        emp_code_in = employee_id
                                        gross_salary_in = gross_earning
                                        emp_gender_in = gender
                                        city_id_in = city_id
                                        state_id_in = state_id
                                        site_id_in = slot.site_id.site_id  


                                        return_value= calculate_professional_tax(user_session, emp_code_in, act_id_in, re_year_in, re_month_in, state_id_in, city_id_in, site_id_in, gross_salary_in,emp_gender_in,comp_id_in)

                                        if working_hours < 9:
                                            amount = return_value
                                        else:
                                            amount = return_value


                                    elif element.item_name == 'LWF Employer':
                                       
                                    
                                        shift_date = slot.shift_date
                                        re_year_in = shift_date.year
                                        re_month_in = shift_date.month

                                        comp_id_in = slot.company.company_id
                                        act_id_in = 2
                                        user_id_in = user_session
                                        emp_code_in = employee_id
                                        gross_salary_in = gross_earning
                                        emp_gender_in = gender
                                        city_id_in = city_id
                                        state_id_in = state_id
                                        site_id_in = slot.site_id.site_id  

                                        employer_amount = get_employer_deduct(comp_id_in, act_id_in, user_id_in, emp_code_in, city_id_in, state_id_in, site_id_in, re_month_in, re_year_in)


                                        # return_value= calculate_lwf_tax(user_session, emp_code_in, act_id_in, re_year_in, re_month_in, state_id_in, city_id_in, site_id_in, gross_salary_in,emp_gender_in,comp_id_in)

                                        # if working_hours < 9:
                                        #     amount = return_value
                                        # else:
                                        #     amount = return_value
                                       #  Calculate percentage based on BASIC
                                        if working_hours < 9:
                                            amount = employer_amount
                                        else:
                                            amount = employer_amount
                                    elif element.item_name == 'Income Tax':

                                        gross_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            element_name='Gross Earning',  # Add filter for element_name
                                            pay_type='Total Earning'  # Retain pay_type filter if necessary
                                        ).aggregate(
                                            total_gross_earning=Sum('amount')  # Use Sum to aggregate the 'amount' column
                                        )['total_gross_earning'] or 0

                                        re_year_in = shift_date.year
                                        re_month_in = shift_date.month
                                        emp_code_in = employee_id

                                        
                                        income_tax = calculate_income_tax(user_session, emp_code_in, re_year_in, re_month_in,gross_earning)

                                        if working_hours < 9:
                                            amount = income_tax
                                        else:
                                            amount = income_tax

                                    #     try:
                                    #         deduction = income_tax_deduction.objects.get(
                                    #             employee_id=employee_id,
                                    #             company_id=employee.company_id.company_id,
                                    #             deduction_month=attendance.attendance_date.month,
                                    #             deduction_year = attendance.attendance_date.year,
                                    #             is_deducted = 0
                                    #         )
                                    #         amount = deduction.deduction_amount
                                    #     except income_tax_deduction.DoesNotExist:
                                    #         amount = 0
                                    else:
                                        if working_hours < 9:
                                            amount = element.four_hour_amount
                                        else:
                                            amount = element.nine_hour_amount
                                    if element.item_name == 'Gross Earning':
                                        total_earnings = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Earning'
                                        ).exclude(
                                            element_name__in=['Transport Allowance', 'Night Fees']  # Exclude specific element names
                                        ).aggregate(
                                            Sum('amount')
                                        )['amount__sum'] or 0

                                        print(f"Total Gross Earnings: {total_earnings}")

                                        amount = total_earnings 
                                    elif element.item_name == 'Gross Deduction':
                                        total_deduction = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Deduction'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        income_tax = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Income Tax'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        amount = total_deduction + income_tax
                                    
                                    elif element.item_name == 'Net Salary':
                                        total_earnings = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type__in=['Earning', 'Other Earning'] 
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        total_deductions = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Deduction'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        # Calculate net salary as the difference
                                        net_salary = total_earnings - total_deductions
                                        amount = net_salary  # Store net salary in amount
                                        print(net_salary)
                                    elif  element.item_name == 'Employer Total':
                                        # Sum all amounts from daily_salary for this employee, slot, date, and 'earning' pay_type
                                        employer_total = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Employer Contribution'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0
                                        amount = employer_total
                                    elif element.item_name == 'Total Earning':
                                        # Filter for 'Earning' and 'Other Earning' in pay_type
                                        total_earning = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type__in=['Earning', 'Other Earning']  # Include both Earning and Other Earning
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        # Set the calculated total as the amount
                                        amount = total_earning
                                    elif element.item_name == 'CTC':
                                        # Filter for 'Employer Total' and 'Gross Earning' in item_name
                                        total_earn = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Earning'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        employer_total = daily_salary.objects.filter(
                                            employee_id=employee_id,
                                            slot_id=slot_id,
                                            attendance_date=attendance.attendance_date,
                                            pay_type='Employer Contribution'
                                        ).aggregate(Sum('amount'))['amount__sum'] or 0

                                        # If the sum is None, set amount to 0
                                        ctc = total_earn + employer_total
                                        amount = ctc  
                                    if slot.night_shift == 'True':  # Check if the slot is a night shift
                                        if element.item_name == 'Transport Allowance':
                                            # Calculate percentage based on BASIC for Transport Allowance
                                            if working_hours < 9:
                                                amount = element.four_hour_amount
                                            else:
                                                amount = element.nine_hour_amount
                                        elif element.item_name == 'Night Fees':
                                            # Calculate percentage based on BASIC for Night Fees
                                            if working_hours < 9:
                                                amount = element.four_hour_amount
                                            else:
                                                amount = element.nine_hour_amount


                                else:
                                        # Absolute Amount 
                                    if working_hours < 9:
                                        amount = element.four_hour_amount
                                    else:
                                        amount = element.nine_hour_amount
                                
                                try:
                                    # amount = math.ceil(amount)
                                    daily_salary.objects.create(
                                        slot_id=slot,
                                        employee_id=employee_id,
                                        company_id = employee.company_id,
                                        attendance_date=attendance.attendance_date,
                                        tax_parameter = element.tax_parameter,
                                        work_hours=working_hours,
                                        amount=amount,
                                        element_name=element.item_name,
                                        pay_type=element.pay_type,
                                        classification=element.classification,
                                        created_by=request.user,
                                        updated_by=request.user
                                    )
                                except Exception as e:
                                    tb = traceback.format_exc()
                                    print(f"Error: {e}")
                                    
                            salary_generated_log.objects.create(slot_id=slot,
                                    employee_id=employee_id,
                                    company_id = employee.company_id,
                                    slot_date=attendance.attendance_date,
                                    created_by=request.user,
                                    updated_by=request.user
                                    )
                            
                            user_slot = UserSlotDetails.objects.get(slot_id=slot, employee_id=employee_id)

                            if user_slot:
                                status = get_object_or_404(PayrollStatusMaster, status_id=3)  # Assuming you want to update the status to 2
                                user_slot.status = status
                                user_slot.save()

                            slot = SlotDetails.objects.get(slot_id=slot_id)
                            if slot:
                                status = get_object_or_404(PayrollStatusMaster, status_id = 3)
                                slot.status = status
                                slot.save()
                            try:
                                deduction = income_tax_deduction.objects.get(
                                    employee_id=employee_id,
                                    company_id=employee.company_id,
                                    deduction_month=attendance.attendance_date.month,
                                    deduction_year=attendance.attendance_date.year,
                                    is_deducted=False
                                )
                                if deduction:
                                    deduction.is_deducted = 1
                                    deduction.deducted_on = attendance.attendance_date
                                    deduction.save()   
                                # Do something with deduction
                            except income_tax_deduction.DoesNotExist:
                                deduction = None  
                else:
                    # Invalid time format (missing ':')
                    working_hours = 0
            else:
                # Either time_in or time_out is None
                working_hours = 0
    # return redirect('slot_list')
    return redirect(f'/user_salary_index?slot_id={slot_id}')
class SlotListView(ListView):
    model = SlotDetails
    template_name = 'Payroll/Slot/index.html'
    context_object_name = 'slots'
    paginate_by = 50  
    def get_queryset(self):
        # Filter the SlotDetails based on status_id = 2
        return SlotDetails.objects.filter(status_id=2)
    
class ApproveSlotListView(ListView):
    model = SlotDetails
    template_name = 'Payroll/Slot/approve_index.html'
    context_object_name = 'slots'
    paginate_by = 150  

    # def get_queryset(self):
    #     # Exclude SlotDetails with status_id = 1 or 2
    #     return SlotDetails.objects.exclude(status__in=[1, 2])


    
def generate_salary_redirect(request, slot_id):
    return redirect('generate_salary', slot_id=slot_id)

def user_salary_index(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    user = request.session.get('user_id', '')
    try:
        # Get the slot ID from the request
        slot_id = request.GET.get('slot_id', '')
        slot = get_object_or_404(SlotDetails, slot_id=slot_id)

        # Retrieve user slot details for the given slot ID
        user_slot_details = UserSlotDetails.objects.filter(slot_id=slot_id)

        # Calculate total salary by fetching net salary for each user slot
        for user_slot in user_slot_details:
            # Filter the record where element_name is 'net_salary' for the given employee and slot_id
            net_salary_record = daily_salary.objects.filter(
                employee_id=user_slot.employee_id, 
                slot_id=slot_id, 
                element_name='Net Salary'
            ).first()
            
            # Set the values for total_salary and amount based on the filtered record
            user_slot.total_salary = net_salary_record.amount if net_salary_record else 0
            user_slot.amount = net_salary_record.amount if net_salary_record else 0


        # Prepare the context for rendering the template
        context = {
            'slot': slot,
            'user_slot_details': user_slot_details,
        }
        return render(request, 'Payroll/Slot/user_salary_index.html', context)
    except Exception as e:
        print(f"Error: {e}") 
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log", [fun, str(e), user])
        messages.error(request, 'Oops...! Something went wrong!')
    
def view_approve_salary(request,slot_id):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()
    user = request.session.get('user_id', '')
    try:
        slot = get_object_or_404(SlotDetails, slot_id=slot_id)

        # Retrieve user slot details for the given slot ID
        user_slot_details = UserSlotDetails.objects.filter(slot_id=slot_id)

        # Calculate total salary by fetching net salary for each user slot
        for user_slot in user_slot_details:
            # Filter the record where element_name is 'net_salary' for the given employee and slot_id
            net_salary_record = daily_salary.objects.filter(
                employee_id=user_slot.employee_id, 
                slot_id=slot_id, 
                element_name='Net Salary'
            ).first()
            
            # Set the values for total_salary and amount based on the filtered record
            user_slot.total_salary = net_salary_record.amount if net_salary_record else 0
            user_slot.amount = net_salary_record.amount if net_salary_record else 0


        # Prepare the context for rendering the template
        context = {
            'slot': slot,
            'user_slot_details': user_slot_details,
        }
        return render(request, 'Payroll/Slot/view_approve_salary.html', context)

    except Exception as e:
        print(f"Error: {e}") 
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
    
        cursor.callproc("stp_error_log", [fun, str(e), user])
        messages.error(request, 'Oops...! Something went wrong!')
    

@login_required
def user_slot_details_list(request, slot_id):
    slot = get_object_or_404(SlotDetails, slot_id=slot_id)
    user_slot_details = slot_attendance_details.objects.filter(slot_id=slot_id)

    context = {
        'slot': slot,
        'user_slot_details': user_slot_details,
    }
    return render(request, 'Payroll/Slot/user_slot_details.html', context)

@login_required
def view_employee_salary_details(request, employee_id, slot_id):
    try:
        # Retrieve the slot object
        slot = get_object_or_404(SlotDetails, slot_id=slot_id)
        date = slot.shift_date

        # Retrieve user slot details for the given slot ID
        user_slot_details = UserSlotDetails.objects.filter(slot_id=slot_id).first()
        status_id = user_slot_details.status.status_id if user_slot_details else None  # Handle case when no data found

        if user_slot_details:
    # Fetch the employee record based on employee_id
            employee = sc_employee_master.objects.filter(employee_id=employee_id).first()
            if employee:
                employee_name = employee.employee_name
            else:
                employee_name = None  # Handle case where employee is not found
        else:
            employee_name = None 

        # Filter daily_salary records for the given slot_id and employee_id
        daily_salary_data = daily_salary.objects.filter(slot_id=slot_id, employee_id=employee_id)

        # Separate daily_salary data into two lists: earnings and deductions
        earnings = daily_salary_data.filter(pay_type__in=['Earning','Total Earning','Other Earning'])
        deductions = daily_salary_data.filter(pay_type__in=['Deduction','Income Tax'])
        employer_contribution = daily_salary_data.filter(pay_type__in=['Employer Contribution'])
        employer_contribution_total = daily_salary_data.filter(pay_type__in=['Employer Total'])

        # Calculate the total earnings (Gross Earning and Net Salary)
        # Fetch rows for earnings
        total_earnings_value = daily_salary_data.filter(
            # pay_type__in='Total',
            element_name__in=['Total Earning', 'Net Salary']
        )

        # Fetch rows for deductions
        total_deductions_value = daily_salary_data.filter(
            pay_type='Total',
            element_name__in=['Gross Deduction', 'Net Salary']
        )


        # Prepare the context for rendering the template
        context = {
            'slot_id':slot_id,
            'status_id':status_id,
            'date':date,
            'employee_id':employee_id,
            'employee_name':employee_name,
            'earnings': earnings,  # Earnings list
            'deductions': deductions,  # Deductions list
            'total_earnings': earnings,  # Earnings used for individual elements
            'total_deductions': deductions,  # Deductions used for individual elements
            'total_earnings_value': total_earnings_value,  # Total Earnings value
            'total_deductions_value': total_deductions_value,
            'employer_contribution':employer_contribution,
            'employer_contribution_total':employer_contribution_total 
        }

        # Render the response with context
        return render(request, 'Payroll/Slot/view_employee_salary_details.html', context)

    except Exception as e:
        print(f"Error: {e}")
        messages.error(request, 'Oops...! Something went wrong!')
        return render(request, 'Payroll/Slot/view_employee_salary_details.html', {'error': str(e)})



@login_required
def handle_card_name_change(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)  # Safely parse JSON data
            card_name = data.get('card_name')
            
            if not card_name:
                return JsonResponse({"error": "Card name is required."}, status=400)

            # Check if the card_name already exists in the database
            if rate_card_master.objects.filter(card_name=card_name).exists():
                return JsonResponse({"error": "Card name already exists!"}, status=400)
            else:
                # You can perform further actions here if needed (e.g., saving the card name)
                return JsonResponse({"message": "Card name is available."})

        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data."}, status=400)

    return JsonResponse({"error": "Invalid request method."}, status=405)

@login_required
def download_sample(request):
    user = request.session.get('user_id', '')

    try:
       
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Data"

        # Define NamedStyles for headers and data rows
        header_style = NamedStyle(name="header_style", font=Font(bold=True, size=14))
        data_style = NamedStyle(name="data_style", font=Font(size=10))

        # Add headers for attendance data
        header_row = ["Attendance Date", "Employee Id", "Attendance In", "Attendance Out"]
        ws.append(header_row)
        for cell in ws[1]:  # Apply header style to the first row
            cell.style = header_style

        # Sample attendance data
        time_format = "%H:%M"  # 24-hour format with hours and minutes

        attendance_data = [
            ["2024-11-05", "PSNID01", 
            datetime.strptime("09:00:00", "%H:%M:%S").strftime(time_format), 
            datetime.strptime("17:00:00", "%H:%M:%S").strftime(time_format)],
            ["2024-11-05", "PSNID02", 
            datetime.strptime("09:30:00", "%H:%M:%S").strftime(time_format), 
            datetime.strptime("17:30:00", "%H:%M:%S").strftime(time_format)],
        ]

        # Add attendance data to the worksheet
        for row in attendance_data:
            ws.append(row)

        # Apply the data style to the data rows (starting from the 2nd row)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.style = data_style

        # Create a response object for downloading the Excel file
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=attendance_data.xlsx"

        # Save the workbook to the response
        wb.save(response)

        return response

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        Cursor.callproc("stp_error_log", [fun, str(e), request.session.get('user_id', '')])
        messages.error(request, 'Oops...! Something went wrong!')


def attendance_error(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    pre_url = request.META.get('HTTP_REFERER')
    header, data = [], []
    global user
    user  = request.session.get('user_id', '')
    try:
         
       if request.method=="GET":
            entity = 'att'
            type = 'err'
            cursor.callproc("stp_get_masters",[entity,type,'name',user])
            for result in cursor.stored_results():
                datalist1 = list(result.fetchall())
            name = datalist1[0][0]
            cursor.callproc("stp_get_masters", [entity, type, 'header',user])
            for result in cursor.stored_results():
                header = list(result.fetchall())
            cursor.callproc("stp_get_masters",[entity,type,'data',user])
            for result in cursor.stored_results():
                data = list(result.fetchall())


    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        if request.method=="GET":
            return render(request,'Master/index.html', {'type':type,'header':header,'name':name,'data':data,'pre_url':pre_url})
        

def approve_attendance(request):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        if request.method == "POST":
            try:
                slot_id = request.POST.get('slot_id')
                slot_name = request.POST.get('slot_name')

                # Fetch all records in SlotAttendanceDetails for the given slot_id
                slot_attendance_records = slot_attendance_details.objects.filter(slot_id=slot_id)
                if slot_attendance_records.exists():
                # Update the status for all matching SlotAttendanceDetails records
                    status = get_object_or_404(PayrollStatusMaster, status_id=2)
                    slot_attendance_records.update(status=status)

                    # Update the status for UserSlotDetails based on the employee_id from SlotAttendanceDetails
                    for record in slot_attendance_records:
                        # Fetch UserSlotDetails where slot_id matches and employee_id matches the one in the SlotAttendance record
                        user_slot = UserSlotDetails.objects.filter(slot_id=slot_id, employee_id=record.employee_id).first()
                        if user_slot:
                            # Update the status for each matching UserSlotDetails record
                            user_slot.status = get_object_or_404(PayrollStatusMaster, status_id=2)
                            user_slot.save()

                # Update the status for the corresponding SlotDetails record
                slot_details = get_object_or_404(SlotDetails, slot_id=slot_id)
                slot_details.status = get_object_or_404(PayrollStatusMaster, status_id=2)
                slot_details.save()  # Save the changes


                # Show success message
                messages.success(request, "Attendance successfully approved for Slot : {}".format(slot_name))
                return redirect( f'/attendance_index?type=index')
            except Exception as e:
                # Show error message if something goes wrong
                messages.error(request, "An error occurred: {}".format(str(e)))
            
        else:
            slot_data = SlotDetails.objects.filter(slot_id__in=slot_attendance_details.objects.values('slot_id')).all() 
            return render(request, 'Payroll/Attendance/approve.html', {'slot_data': slot_data})
    
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()


def edit_attendance(request , encrypted_id):
    Db.closeConnection()
    m = Db.get_connection()
    cursor=m.cursor()
    global user
    user  = request.session.get('user_id', '')
    try:
        if request.method == "GET":
            id = dec(encrypted_id)
            slot_attendance = get_object_or_404(slot_attendance_details, id=id)
            # print(slot_attendance.status)
            status_values = PayrollStatusMaster.objects.filter(status_type='Attendance').values('status_id', 'status_name')



            # Step 2: Retrieve the employee details based on the employee_id from SlotAttendanceDetails
            employee_id = slot_attendance.employee_id
            employee_details = get_object_or_404(sc_employee_master, employee_id=employee_id)

            # Step 3: Pass the data to the template or handle your logic
            context = {
                'slot_attendance': slot_attendance,
                'employee_details': employee_details,
                'status_values':status_values
            }

            return render(request, 'Payroll/Attendance/edit_attendance.html', context)
        if request.method == "POST":

            id = dec(encrypted_id)
            # Step 2: Retrieve the corresponding slot attendance details using the id
            slot_attendance = get_object_or_404(slot_attendance_details, id=id)
            employee_name = request.POST.get('employee_name', '')
            status = request.POST.get('status', '')

            # Step 3: Retrieve the attendance_in and attendance_out from the form (POST request)
            attendance_in = request.POST.get('attendance_in')  # Ensure this field is in your form
            attendance_out = request.POST.get('attendance_out')  # Ensure this field is in your form

            # Step 4: Update the attendance_in and attendance_out fields
            slot_attendance.attendance_in = attendance_in
            slot_attendance.attendance_out = attendance_out
            slot_attendance.status = get_object_or_404(PayrollStatusMaster, status_id = status)

            # Step 5: Save the updated record
            slot_attendance.save()
            messages.success(request, f'Attendance for {employee_name} got updated successfully!')
            return redirect('/attendance_index?type=index')
        else:
            messages.error(request, 'Error updating Attendance.')

        
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])  
        messages.error(request, 'Oops...! Something went wrong!')
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()



def create_payout(request):
    user  = request.session.get('user_id', '')
    try:
        if request.method == 'POST':
        # Get the list of employee_ids as strings
            employee_ids = request.POST.getlist('employee_ids')
            print("Employee IDs: ", employee_ids)  # Print employee_ids for debugging
            slot_id = int(request.POST.get('slot_id', 0))  # Ensure slot_id is an integer

            try:
                # Try fetching SlotDetails for the given slot_id
                slot_details = SlotDetails.objects.get(slot_id=slot_id)
            except SlotDetails.DoesNotExist:
                return redirect('error_page')

            for employee_id in employee_ids:  # Loop through each employee_id (still a string)
                print(f"Processing Employee ID: {employee_id}") 
                
                if PayoutDetails.objects.filter(employee_id=employee_id, slot_id=slot_id).exists():
                    print(f"Payout already exists for Employee ID: {employee_id} and Slot ID: {slot_id}")
                    continue 
                
                employee = sc_employee_master.objects.get(employee_id=employee_id)
                payment_details = pay.objects.first()
                # Fetch salary details
                salary = daily_salary.objects.filter(element_name="net salary", employee_id=employee_id).values_list('amount', flat=True).first()
                if salary is None:
                    raise ValueError("No net salary found for the given employee.")
                # Prepare the API payload
                payload = {
                    "account_number": payment_details.account_number,
                    "amount": int(salary),
                    "currency": payment_details.currency,
                    "mode": payment_details.mode,
                    "purpose": payment_details.purpose,
                    "fund_account": {
                        "account_type": "bank_account",
                        "bank_account": {
                            "name": employee.account_holder_name,
                            "ifsc": employee.ifsc_code,
                            "account_number": employee.account_no,
                        },
                        "contact": {
                            "name": employee.employee_name,
                            "email": employee.email,
                            "contact": employee.mobile_no,
                            "type": "employee",
                            "reference_id": str(employee.id),
                        },
                    },
                    "queue_if_low_balance": True,
                    "reference_id": str(employee.id),
                    "narration": "Acme Corp Fund Transfer",
                }
                # Create payout details
                payout = PayoutDetails.objects.create(
                    amount=salary,
                    company_id=get_object_or_404(company_master, company_id=slot_details.company.company_id),
                    site_id=get_object_or_404(site_master, site_id=slot_details.site_id.site_id),
                    employee_id=employee_id,
                    slot_id=get_object_or_404(SlotDetails, slot_id=slot_details.slot_id),
                    payout_status=get_object_or_404(PayrollStatusMaster, status_id=5),
                    payment_initiated_date=date.today(),
                    created_by=get_object_or_404(CustomUser, id=user),
                    updated_by=get_object_or_404(CustomUser, id=user),
                )
                slot_id = slot_details.slot_id  # Assuming `slot_details` has been fetched already
                SlotDetails.objects.filter(slot_id=slot_id).update(
                    status=get_object_or_404(PayrollStatusMaster, status_id=5)  # Replace '6' with the desired status ID
                )
                # Call Razorpay API to create payout
                response = requests.post(
                    url="https://api.razorpay.com/v1/payouts",
                    json=payload,
                    auth=(payment_details.api_key, payment_details.secret_key),
                )
                if response.status_code == 200 or response.status_code == 201:
                    print(f"Payout successful for Employee ID: {employee_id}")
                    response_data = response.json()  # Parse the JSON response
                    razor_payout_id = response_data.get('id')
                    fund_account_id = response_data.get('fund_account_id')
                    reference_id = response_data.get('reference_id')
                    purpose = response_data.get('purpose')
                    status = response_data.get('status')
                    tax = response_data.get('tax')
                    fees = response_data.get('fees')
                    if fund_account_id and status:
                        payout.razorpay_payout_id = razor_payout_id
                        payout.fund_account_id = fund_account_id
                        payout.fees = fees
                        payout.tax = tax
                        payout.reference_id = reference_id
                        payout.purpose = purpose
                        payout.payout_for_date = date.today()
                        if status == 'processing':
                            payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=6)
                            slot_id = slot_details.slot_id  
                            UserSlotDetails.objects.filter(slot_id=slot_id, employee_id=employee_id).update(
                                status=get_object_or_404(PayrollStatusMaster, status_id=6)
                            )
                            # SlotDetails.objects.filter(slot_id=slot_id).update(
                            #     status=get_object_or_404(PayrollStatusMaster, status_id=6)  
                            # )
                        payout.save()  # Save the updated record
                        print(f"Payout updated for Employee ID: {employee_id}, Fund Account ID: {fund_account_id}, Status: {status}")
                    else:
                        print(f"Missing fund_account_id or status in Razorpay response for Employee ID: {employee_id}")
                else:
                    response_data = response.json()
                    reason = response_data.get('error', {}).get('description', 'Unknown error')
                    
                    payout.failure_reason = reason
                    payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=9)
                    payout.save()
                    employee_id = payout.employee_id
                    
                    UserSlotDetails.objects.filter(slot_id=slot_id, employee_id=employee_id).update(
                        status=get_object_or_404(PayrollStatusMaster, status_id=9)
                    )

        messages.success(request, 'Salary Successfully Started Processing to the Employees')
        return redirect('approveslots')
    
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        messages.error(request, 'Oops...! Something went wrong!')


class UpdatePayoutStatus(APIView):

    def get(self, request):
            errors = []
            success = []
        
            all_payouts = PayoutDetails.objects.all()
            payment_details = pay.objects.first()
            if not payment_details:
                return response({"error": "Payment details not found."}, status=400)

            api_key = payment_details.api_key
            secret_key = payment_details.secret_key


            for payout in all_payouts:
                try:
                    payout_id = payout.razorpay_payout_id
                    slot = payout.slot_id.slot_id
                    # employee = payout.employee_id

                    response = None

                    # Fetch payout details from Razorpay API using Basic Auth
                    response = requests.get(
                        f"https://api.razorpay.com/v1/payouts/{payout_id}",
                        auth=(api_key, secret_key)
                    )

                    if response.status_code == 200:
                        response_data = response.json()
                        status = response_data.get('status')
                        utr = response_data.get('utr') 
                        failure_reason = response_data.get('failure_reason', 'Unknown reason')

                        # Update PayoutDetails entry based on the status
                        if status == 'processed':
                            payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=7)
                            payout.utr = utr 
                            slot_details = SlotDetails.objects.filter(slot_id=slot)  # Fetch all matching SlotDetails entries
                            if slot_details.exists():  # Check if any entries exist
                                for slot in slot_details:  # Iterate through each matching entry
                                    slot.status = get_object_or_404(PayrollStatusMaster, status_id=7)  
                                    slot.save()
                                    # send_salary_notification(slot, payout.employee_id, payout.amount)
                        elif status == 'processing':
                            payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=6)  
                            slot_details = SlotDetails.objects.filter(slot_id=slot)
                            payout.utr = None
                            if slot_details.exists():  # Check if any entries exist
                                for slot in slot_details:  # Iterate through each matching entry
                                    slot.status = get_object_or_404(PayrollStatusMaster, status_id=6)  
                                    slot.save()
                        elif status == 'reversed':
                            payout.failure_reason = failure_reason
                            payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=8)
                            slot_details = SlotDetails.objects.filter(slot_id=slot)
                            payout.utr = None
                            if slot_details.exists():  # Check if any entries exist
                                for slot in slot_details:  # Iterate through each matching entry
                                    slot.status = get_object_or_404(PayrollStatusMaster, status_id=8)  
                                    slot.save()
                        payout.payment_completed_date = datetime.now()
                        payout.save()

                        success.append(f"Successfully Payment Sent to ", payout.employee_id)
                    else:
                        errors.append(f"Error Sending payment to ", payout.employee_id)
                except Exception as e:
                    print(f"Error updating payout {payout_id}: {str(e)}")
                    if response is not None:
                        print(f"HTTP Response: {response.text}")



# def refresh_payout_status(request):
#     errors = []
#     success = []
    
#     slot = request.GET.get('slot_id')
    
#     if not slot:
#         return JsonResponse({"error": "Slot ID not provided."}, status=400)

#     all_payouts = PayoutDetails.objects.filter(slot_id=slot)
#     payment_details = pay.objects.first()

#     if not payment_details:
#         return JsonResponse({"error": "Payment details not found."}, status=400)

#     api_key = payment_details.api_key
#     secret_key = payment_details.secret_key

#     for payout in all_payouts:
#         try:
#             payout_id = payout.razorpay_payout_id
#             slot = payout.slot_id.slot_id
#             response = None

#             # Fetch payout details from Razorpay API using Basic Auth
#             response = requests.get(
#                 f"https://api.razorpay.com/v1/payouts/{payout_id}",
#                 auth=(api_key, secret_key)
#             )

#             if response.status_code == 200:
#                 response_data = response.json()
#                 status = response_data.get('status')
#                 utr = response_data.get('utr') 
#                 failure_reason = response_data.get('failure_reason', 'Unknown reason')

#                 # Update PayoutDetails entry based on the status
#                 if status == 'processed':
#                     payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=7)
#                     payout.utr = utr 
#                     slot_details = SlotDetails.objects.filter(slot_id=slot)  # Fetch all matching SlotDetails entries
#                     if slot_details.exists():  # Check if any entries exist
#                         for slot in slot_details:  # Iterate through each matching entry
#                             slot.status = get_object_or_404(PayrollStatusMaster, status_id=7)  
#                             slot.save()
#                 elif status == 'processing':
#                     payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=6)  
#                     slot_details = SlotDetails.objects.filter(slot_id=slot)
#                     payout.utr = None
#                     if slot_details.exists():  # Check if any entries exist
#                         for slot in slot_details:  # Iterate through each matching entry
#                             slot.status = get_object_or_404(PayrollStatusMaster, status_id=6)  
#                             slot.save()
#                 elif status == 'reversed':
#                     payout.failure_reason = failure_reason
#                     payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=8)
#                     slot_details = SlotDetails.objects.filter(slot_id=slot)
#                     payout.utr = None
#                     if slot_details.exists():  # Check if any entries exist
#                         for slot in slot_details:  # Iterate through each matching entry
#                             slot.status = get_object_or_404(PayrollStatusMaster, status_id=8)  
#                             slot.save()
#                 payout.payment_completed_date = datetime.now()
#                 payout.save()

#                 success.append(f"Successfully updated payout {payout_id}")
#             else:
#                 errors.append(f"Error fetching payout details for {payout_id}: {response.text}")

#         except Exception as e:
#             errors.append(f"Error updating payout {payout_id}: {str(e)}")
#             if response is not None:
#                 errors.append(f"HTTP Response: {response.text}")

#     # After processing all payouts, show success or error messages
#     if errors:
#         return JsonResponse({"success": False, "errors": "Error"})
#     return JsonResponse({"success": True, "messages": "Success"})



def generate_pay_slip(request):
    try:
        # Get parameters from the GET request
        employee_id = request.GET.get('employeeId')
        slot_id = request.GET.get('slotId')
        image_path = 'static/images/psn-logo0.png'
        user = request.session.get('user_id', '')

        # Open the image using Pillow
        with Image.open(image_path) as img:
            # Convert image to BytesIO object
            buffer = BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)

            # Encode the image to base64
            encoded_image = base64.b64encode(buffer.read()).decode('utf-8')

        if not employee_id or not slot_id:
            return HttpResponse("Missing employeeId or slotId", status=400)

        # Fetch salary details and employee data
        salary_details = daily_salary.objects.filter(employee_id=employee_id, slot_id=slot_id)
        employee = get_object_or_404(UserSlotDetails, employee_id=employee_id, slot_id=slot_id)
        slot_attended = get_object_or_404(slot_attendance_details, employee_id=employee_id, slot_id=slot_id)

        # Calculate the time difference
        attendance_in_str = slot_attended.attendance_in
        attendance_out_str = slot_attended.attendance_out

        # Calculate the difference between attendance_in and attendance_out
        time_format = "%H:%M:%S"  # 24-hour format (HH:MM)

        # Convert to datetime objects (use the same date for both to compare times)
        attendance_in = datetime.strptime(attendance_in_str, time_format)
        attendance_out = datetime.strptime(attendance_out_str, time_format)

        # Calculate the time difference
        time_difference = attendance_out - attendance_in

        # Check if the difference is more than or equal to 9 hours
        if time_difference >= timedelta(hours=9):
            result = f"{attendance_in.strftime(time_format)} - {attendance_out.strftime(time_format)} (9 hours )"
        else:
            result = f"{attendance_in.strftime(time_format)} - {attendance_out.strftime(time_format)} (4 hours )"

        # The result will now contain the time difference along with the "9 hours" or "4 hours" part
        print(result)
        slot = get_object_or_404(SlotDetails, slot_id=slot_id)
        
        # Filter earnings and deductions
        earnings = salary_details.filter(pay_type__in=['Earning', 'Other Earning', 'Total Earning'])
        # deductions = salary_details.filter(Q(pay_type='Deduction') | Q(element_name='Gross Deduction'))
        deductions = salary_details.filter(Q(pay_type='Deduction') | Q(element_name='Income Tax') | Q(element_name='Gross Deduction'))


        total_earnings = salary_details.filter(pay_type='Total', element_name='Total Earning').first()
        gross_earnings = salary_details.filter(pay_type='Total Earning', element_name='Gross Earning').first()
        net_salary = salary_details.filter(pay_type='Total', element_name='Net Salary').first()
        gross_deductions = salary_details.filter(pay_type='Total', element_name='Gross Deduction').first()

        date_today = datetime.now().strftime('%d-%m-%Y')
        shift_date = slot.shift_date.strftime('%d-%m-%Y')

        context = {
            'result': result,
            'slot': slot,
            'shift_date': shift_date,
            'employee_name': employee.emp_id.employee_name,  # Assuming `emp_id` is a ForeignKey with the employee details
            'employee_id': employee_id,
            'earnings': [{'name': e.element_name, 'amount': e.amount} for e in earnings],
            'deductions': [{'name': d.element_name, 'amount': d.amount} for d in deductions],
            'gross_earnings': gross_earnings.amount,
            'net_salary': net_salary.amount,
            'gross_deductions': gross_deductions.amount,
            'total_earnings_amount': total_earnings.amount,
            'date': date_today,
            'encoded_image': encoded_image
        }

        # Render the template to HTML
        html_content = render_to_string('Payroll/Slot/payment_slip.html', context)

        # Create a PDF buffer (in-memory storage for the PDF)
        pdf_buffer = BytesIO()

        # Generate the PDF and write it to the buffer
        pisa_status = pisa.CreatePDF(html_content, dest=pdf_buffer)

        # Check for errors in PDF generation
        if pisa_status.err:
            return HttpResponse(f"Error generating PDF: {pisa_status.err}", status=500)
        
        file_name = f'payslip_{employee_id}_{slot_id}.pdf'

        # Construct the full file path using os.path.join
        file_path = os.path.join(settings.MEDIA_ROOT, file_name)

        # Ensure the directory exists (optional but recommended)
        if not os.path.exists(os.path.dirname(file_path)):
            os.makedirs(os.path.dirname(file_path))

        # Save the PDF file to the specified location
        with open(file_path, 'wb') as pdf_file:
            pdf_file.write(pdf_buffer.getvalue())

        # Save pay slip record to the database (if you want to track the file path)
        existing_pay_slip = PaySlip.objects.filter(employee_id=employee_id, slot_id=slot_id).first()

        if existing_pay_slip:
            # If the record exists, update the existing PaySlip
            existing_pay_slip.pdf_file = file_path
            existing_pay_slip.created_by = get_object_or_404(CustomUser, id = user) 
            existing_pay_slip.save()
        else:
            # If no record exists, create a new PaySlip
            pay_slip = PaySlip(
                employee_id=employee_id,
                slot_id=get_object_or_404(SlotDetails,slot_id =  slot_id),
                pdf_file=file_path ,
                created_by = get_object_or_404(CustomUser, id = user) # Save the file path in the model
            )
            pay_slip.save()

        # Create a response to download the PDF
        response = HttpResponse(open(file_path, 'rb').read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payslip_{employee_id}_{slot_id}.pdf"'
        return response

    except Exception as e:
        # Capture and print error traceback for debugging
        tb = traceback.format_exc()
        print(f"Error: {e}")
        print(f"Traceback: {tb}")
        
        # Return error message in response
        return HttpResponse(f"Something went wrong: {str(e)}", status=500)
    

def create_new_payout(request, employee_id, slot_id):
    user = request.session.get('user_id', '')
    try:
        try:
            slot_details = SlotDetails.objects.get(slot_id=slot_id)
        except SlotDetails.DoesNotExist:
            return redirect('error_page')
        print(f"Processing Employee ID: {employee_id}")  # Debugging output
        # Fetch employee details
        employee = sc_employee_master.objects.get(employee_id=employee_id)
        payment_details = pay.objects.first()
        
        # Fetch salary details
        salary = daily_salary.objects.filter(element_name="net salary", employee_id=employee_id).values_list('amount', flat=True).first()
        if salary is None:
            raise ValueError("No net salary found for the given employee.")
        # Prepare the API payload
        payload = {
            "account_number": payment_details.account_number,
            "amount": int(salary),
            "currency": payment_details.currency,
            "mode": payment_details.mode,
            "purpose": payment_details.purpose,
            "fund_account": {
                "account_type": "bank_account",
                "bank_account": {
                    "name": employee.account_holder_name,
                    "ifsc": employee.ifsc_code,
                    "account_number": employee.account_no,
                },
                "contact": {
                    "name": employee.employee_name,
                    "email": employee.email,
                    "contact": employee.mobile_no,
                    "type": "employee",
                    "reference_id": str(employee.id),
                },
            },
            "queue_if_low_balance": True,
            "reference_id": str(employee.id),
            "narration": "Acme Corp Fund Transfer",
        }
        # Create payout details
        payout = PayoutDetails.objects.create(
            amount=salary,
            company_id=get_object_or_404(company_master, company_id=slot_details.company.company_id),
            site_id=get_object_or_404(site_master, site_id=slot_details.site_id.site_id),
            employee_id=employee_id,
            slot_id=get_object_or_404(SlotDetails, slot_id=slot_details.slot_id),
            payout_status=get_object_or_404(PayrollStatusMaster, status_id=5),
            payment_initiated_date=date.today(),
            created_by=get_object_or_404(CustomUser, id=user),
            updated_by=get_object_or_404(CustomUser, id=user),
        )
        # Call Razorpay API to create payout
        response = requests.post(
            url="https://api.razorpay.com/v1/payouts",
            json=payload,
            auth=(payment_details.api_key, payment_details.secret_key),
        )
        if response.status_code in [200, 201]:
            print(f"Payout successful for Employee ID: {employee_id}")
            response_data = response.json()
            razor_payout_id = response_data.get('id')
            fund_account_id = response_data.get('fund_account_id')
            reference_id = response_data.get('reference_id')
            purpose = response_data.get('purpose')
            status = response_data.get('status')
            tax = response_data.get('tax')
            fees = response_data.get('fees')
            if fund_account_id and status:
                payout.razorpay_payout_id = razor_payout_id
                payout.fund_account_id = fund_account_id
                payout.fees = fees
                payout.tax = tax
                payout.reference_id = reference_id
                payout.purpose = purpose
                payout.payout_for_date = date.today()
                if status == 'processing':
                    payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=6)
                    UserSlotDetails.objects.filter(slot_id=slot_id, employee_id=employee_id).update(
                        status=get_object_or_404(PayrollStatusMaster, status_id=6)
                    )
                payout.save()
                messages.success(request, f"Salary Successfully Started Processing for {employee.employee_name}")
            else:
                print(f"Missing fund_account_id or status in Razorpay response for Employee ID: {employee_id}")
        else:
            response_data = response.json()
            reason = response_data.get('error', {}).get('description', 'Unknown error')
            
            payout.failure_reason = reason
            payout.payout_status = get_object_or_404(PayrollStatusMaster, status_id=9)
            payout.save()
            
            UserSlotDetails.objects.filter(slot_id=slot_id, employee_id=employee_id).update(
                status=get_object_or_404(PayrollStatusMaster, status_id=9)
            )
            messages.error(request, f"Error In Salary :{payout.failure_reason}")
        return redirect('approveslots')

    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        messages.error(request, 'Oops...! Something went wrong!')



class payment_details(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, *args, **kwargs):
        try:
            employee_id = request.data.get('employee_id')

            if not employee_id:
                return Response({"error": "employee_id and company_id are required"}, status=400)

            # Filter data by employee_id
            payment_data = salary_generated_log.objects.filter(employee_id=employee_id)

            # Serialize the data
            serialized_data = SalaryGeneratedSerializer(payment_data, many=True).data

            return Response(serialized_data)  # Use Response here instead of response()
        except Exception as e:
            print(e)
            return requests.Response({"error": str(e)}, status=500)
        

class payment_slip_details(APIView):
    permission_classes = [IsAuthenticated]
    authentication_classes = [JWTAuthentication]

    def get(self, request, *args, **kwargs):
        try:
            # Retrieve parameters from request data
            employee_id = request.data.get('employee_id')
            slot_id = request.data.get('slot_id')

            # Validate input parameters
            if not employee_id or not slot_id:
                return Response({"error": "employee_id and slot_id are required"}, status=400)

            # Fetch salary, employee, and attendance details
            salary_details = daily_salary.objects.filter(employee_id=employee_id, slot_id=slot_id)
            employee = get_object_or_404(UserSlotDetails, employee_id=employee_id, slot_id=slot_id)
            slot_attended = get_object_or_404(slot_attendance_details, employee_id=employee_id, slot_id=slot_id)

            # Validate attendance times
            attendance_in_str = slot_attended.attendance_in
            attendance_out_str = slot_attended.attendance_out
            if not attendance_in_str or not attendance_out_str:
                return Response({"error": "Invalid attendance time data"}, status=400)

            # Parse attendance times
            time_format = "%H:%M:%S"
            try:
                attendance_in = datetime.strptime(attendance_in_str, time_format)
                attendance_out = datetime.strptime(attendance_out_str, time_format)
            except ValueError as e:
                return Response({"error": f"Invalid time format: {str(e)}"}, status=400)

            # Calculate time difference
            time_difference = attendance_out - attendance_in
            if time_difference >= timedelta(hours=9):
                result = f"{attendance_in.strftime(time_format)} - {attendance_out.strftime(time_format)} (9 hours)"
            else:
                result = f"{attendance_in.strftime(time_format)} - {attendance_out.strftime(time_format)} (Less than 9 hours)"

            # Fetch slot details
            slot = get_object_or_404(SlotDetails, slot_id=slot_id)

            # Prepare all salary elements in a single list
            salary_elements = [
                {'type': sd.pay_type, 'name': sd.element_name, 'amount': sd.amount}
                for sd in salary_details
            ]

            # Generate response
            context = {
                'result': result,
                'slot': {'slot_name': slot.slot_name, 'shift_date': slot.shift_date.strftime('%d-%m-%Y')},
                'employee_name': employee.emp_id.employee_name if employee.emp_id else "N/A",
                'employee_id': employee_id,
                'salary_elements': salary_elements,
                'date_generated': datetime.now().strftime('%d-%m-%Y'),
            }

            return Response(context)

        except Exception as e:
            print(f"Error: {str(e)}")  # Log error details
            return Response({"error": "Internal server error. Please try again later."}, status=500)

def pt_calculation(user_id_in, comp_id_in, re_year_in, re_month_in, act_id_id_in,emp_gender_in ,emp_code_in, gross_salary_in, city_id_id_in, state_id_id_in, site_id_in):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()

    try:

        user = user_id_in if user_id_in is not None else None
        comp_id = comp_id_in if comp_id_in is not None else None
        re_year = re_year_in if re_year_in is not None else None
        re_month = re_month_in if re_month_in is not None else None
        act_id_id = act_id_id_in if act_id_id_in is not None else None
        emp_code = emp_code_in if emp_code_in is not None else None
        gross_salary = gross_salary_in if gross_salary_in is not None else None
        emp_gender = emp_gender_in if emp_gender_in is not None else None
        city_id_id = city_id_id_in if city_id_id_in is not None else None
        state_id_id = state_id_id_in if state_id_id_in is not None else None

        slab_type = "Employee"  # this function only calculates for PT EMPLOYEE
        salary_period = "" # declared variable             

# -----------  Out Parameters ---------------
      
        comp_id_out = comp_id_in
        state_id_out = state_id_id_in
        city_id_out = city_id_id_in
        act_id_out = act_id_id_in        
        site_id_out = site_id_in                       
        re_year_out = re_year_in
        re_month_out = re_month_in
        emp_code_out = emp_code_in
        gross_salary_out = gross_salary        
                
                                                         
        parameters2 = [state_id_id, act_id_id, city_id_id]
        cursor.callproc("stp_getCheckSlabForStateCity", parameters2)
        for result2 in cursor.stored_results():
            rows2 = result2.fetchall()
            for row3 in rows2:  # slab frequency & month for loop
                slab_freq = row3[0]
                slab_months = row3[1]
                slab_id = row3[2]
                slab_months_challan = row3[3]
                exception_month = row3[4]
                  

# ----------------------- slab not applicable (is_slab is No)---------------------------                                                                                                    
                if (slab_freq == "NA" or slab_months == "NA" or slab_months_challan == "NA"):      
                    salary_period_out = None
                    pt_actual_out = 0
# ---------------------- slab does not exist ----------------------------               
                if (slab_freq == "No" or slab_months == "No" or slab_months_challan == "No"):  
                    salary_period_out = None
                    pt_actual_out = 0
# ---------------------- slab exist (is_slab is Yes)----------------------------                 
                if slab_freq not in ["No","NA"] and slab_months not in ["No","NA"] and slab_months_challan not in ["No","NA"]:

                    if exception_month:
                        # Split the exception_month string by comma to get individual months
                        exception_months = exception_month.split(',')
                        
                        # Initialize status as 'all'
                        slab_applicable_status = 'all'
                        
                        for idx, ex_month in enumerate(exception_months):
                            ex_month1 = int(ex_month)                             
                            if ex_month1 == re_month:
                                # Assign different statuses based on the index
                                if idx == 0:
                                    slab_applicable_status = 'exception'
                                else:
                                    slab_applicable_status = 'exception2'
                                break  # Exit the loop once a match is found
                    else:
                        # If exception_month is null or empty
                        slab_applicable_status = 'all'

                    if slab_freq == "Monthly":
                        checkmonth_formonthly=int(re_month)                                         
                        parameters3 = [slab_id,checkmonth_formonthly,re_year,emp_gender,slab_type,gross_salary,slab_applicable_status]
                        cursor.callproc("stp_getSlabForEmployee", parameters3)
                        for result23 in cursor.stored_results():
                            my = list(result23.fetchall())
                            pt_actual = my[0][0]
                            statuscheck = my[0][1] # this tells if slab found or not
                            
                            if statuscheck == "yes":
                                                  
                                # GET START AND END PERIOD LOGIC 
                                # checkmonth_str = str(checkmonth_formonthly)                                                                                                
                                # salary_period_out = (checkmonth_str + "-" + re_year)
                                checkmonth_str = str(checkmonth_formonthly)  # Convert checkmonth_formonthly to string
                                re_year_str = str(re_year)  # Convert re_year to string
                                salary_period_out = checkmonth_str + "-" + re_year_str  # Concatenate the strings

                                
                                pt_actual_out = pt_actual

                            if statuscheck == "noslabdefined":
                                salary_period_out = None                                
                                pt_actual_out = 0

                    if slab_freq == "Quarterly":
                        checkmonth=int(re_month)
                        checkyear=int(re_year)
                        slab_terms = slab_months_challan
                        matching_period = None
                        new_list_qu = []
                               
# ---------------------  TO CHECK WHICH SLAB MASTER PERIOD IT BELONGS  ------------------------------------------

                        periods = [period.split(',') for period in slab_months_challan.split(' - ')]
                        for period in periods:
                            period_months = list(map(int, period))
                            
                            if checkmonth in period_months:
                                matching_period = period_months
                                break
                        if matching_period:
                            selected_period = matching_period

# ---------------------  TO ASSIGN YEAR TO THE MONTHS  ------------------------------------------

                        if checkmonth == 1:
                            if 1 in selected_period:
                                
                                month_position_1_qu = selected_period.index(1)
                                
                                # Iterate over the list and assign the year based on the position of the month 1                                                
                                for i1, month_qu in enumerate(selected_period):
                                    if i1 >= month_position_1_qu:
                                        # If the month is at or after the position of 1, assign the current year
                                        new_list_qu.append(f'{month_qu}-{checkyear}')
                                    else:
                                        # If the month is before the position of 1, assign the previous year
                                        new_list_qu.append(f'{month_qu}-{checkyear - 1}')                                                
                                                                                                                                                
                        if checkmonth == 12:
                            if 12 in selected_period:
                                
                                month_position_12_qu = selected_period.index(12)

                                # Iterate over the list and assign the year based on the position of the month 12
                                for i2, month_qu in enumerate(selected_period):
                                    if i2 <= month_position_12_qu:
                                        # If the month is at or before the position of 12, assign the current year
                                        new_list_qu.append(f'{month_qu}-{checkyear}')
                                    else:
                                        # If the month is after the position of 12, assign the next year
                                        new_list_qu.append(f'{month_qu}-{checkyear + 1}')                                                
                                                                            
                        if checkmonth != 1 and checkmonth != 12:
                            if 1 in selected_period:
                                index_of_1 = selected_period.index(1)
                                index_of_checkmonth = selected_period.index(checkmonth)
                                if index_of_checkmonth < index_of_1: # for ex take 11
                                    # Values before 1 have the same year
                                    for month_all in selected_period[:index_of_1]:
                                        new_list_qu.append(f"{month_all}-{checkyear}")
                                    # Values from 1 to the right have year + 1
                                    for month_all in selected_period[index_of_1:]:
                                        new_list_qu.append(f"{month_all}-{checkyear + 1}")
                                elif index_of_checkmonth > index_of_1: #for ex take 2
                                # Values to the right of 1 and 1 itself have the same year
                                    for month_all in selected_period[index_of_1:]:
                                        new_list_qu.append(f"{month_all}-{checkyear}")
                                    # Values to the left of 1 have year - 1
                                    for month_all in selected_period[:index_of_1]:
                                        new_list_qu.append(f"{month_all}-{checkyear - 1}")

                            else:
                                for month_all in selected_period:
                                    new_list_qu.append(f"{month_all}-{checkyear}")

                        ordered_list_qu = []
                        for month_qz1 in selected_period:
                            # Find the corresponding year-month combination for each month in selected_period_hy
                            for item_qz1 in new_list_qu:
                                if item_qz1.startswith(f"{month_qz1}-"):
                                    ordered_list_qu.append(item_qz1)
                                    break                                
                        new_list_qu = ordered_list_qu

                                                                                                               
                        parameters3Q = [slab_id,re_month,re_year,emp_gender,slab_type,gross_salary,slab_applicable_status]
                        cursor.callproc("stp_getSlabForEmployee",parameters3Q)
                        for result23Q in cursor.stored_results():
                            myQ = list(result23Q.fetchall())
                            pt_actualQ = myQ[0][0]
                            statuscheckQ = myQ[0][1]
                            # IF SLAB EXISTS THEN SENDING yes ELSE noslabdefined
                            if (statuscheckQ == "yes"):                                              
                                
                                # GET START AND END PERIOD LOGIC                                                
                                last_period_qu = new_list_qu[-1]
                                first_period_qu = new_list_qu[0]                                 
                                                                                               
                                salary_period2 = (first_period_qu + " to " + last_period_qu)
                                salary_period_out = salary_period2                                
                                pt_actual_out = pt_actualQ                                
                            
                            if (statuscheckQ == "noslabdefined"):                                            
                                salary_period_out = None                                
                                pt_actual_out = 0

                    if slab_freq == "Half Yearly":
                        checkmonth_hy=int(re_month)                                     
                        checkyear_hy=int(re_year)
                        matching_period_hy = None
                        new_list_hy = []
                                    
                                                        
# ---------------------  TO CHECK WHICH SLAB MASTER PERIOD IT BELONGS  ------------------------------------------

                        periods_hy = [period_hy.split(',') for period_hy in slab_months_challan.split(' - ')]
                        for period_hy in periods_hy:
                            period_months_hy = list(map(int, period_hy))
                            
                            if checkmonth_hy in period_months_hy:
                                matching_period_hy = period_months_hy
                                break
                        if matching_period_hy:
                            selected_period_hy = matching_period_hy

# ---------------------  TO ASSIGN YEAR TO THE MONTHS  ------------------------------------------

                        if checkmonth_hy == 1:
                            if 1 in selected_period_hy:
                                month_position_1_hy = selected_period_hy.index(1)
                                
                                # Iterate over the list and assign the year based on the position of the month 1                                                
                                for i, m_var in enumerate(selected_period_hy):
                                    if i >= month_position_1_hy:
                                        # If the month is at or after the position of 1, assign the current year
                                        new_list_hy.append(f'{m_var}-{checkyear_hy}')
                                    else:
                                        # If the month is before the position of 1, assign the previous year
                                        new_list_hy.append(f'{m_var}-{checkyear_hy - 1}')

                        if checkmonth_hy == 12:
                            if 12 in selected_period_hy:
                                month_position_12_hy = selected_period_hy.index(12)

                                # Iterate over the list and assign the year based on the position of the month 12
                                for i, m_var in enumerate(selected_period_hy):
                                    if i <= month_position_12_hy:
                                        # If the month is at or before the position of 12, assign the current year
                                        new_list_hy.append(f'{m_var}-{checkyear_hy}')
                                    else:
                                        # If the month is after the position of 12, assign the next year
                                        new_list_hy.append(f'{m_var}-{checkyear_hy + 1}')
                                                                            
                        if checkmonth_hy != 1 and checkmonth_hy != 12:
                            if 1 in selected_period_hy:
                                index_of_1_hy = selected_period_hy.index(1)
                                index_of_checkmonth_hy = selected_period_hy.index(checkmonth_hy)
                                if index_of_checkmonth_hy < index_of_1_hy: #for ex take 11
                                    # Values before 1 have the same year
                                    for month_hy in selected_period_hy[:index_of_1_hy]:
                                        new_list_hy.append(f"{month_hy}-{checkyear_hy}")
                                    # Values from 1 to the right have year + 1
                                    for month_hy in selected_period_hy[index_of_1_hy:]:
                                        new_list_hy.append(f"{month_hy}-{checkyear_hy + 1}")
                                elif index_of_checkmonth_hy > index_of_1_hy: # for ex take 2
                                # Values to the right of 1 and 1 itself have the same year
                                    for month_hy in selected_period_hy[index_of_1_hy:]:
                                        new_list_hy.append(f"{month_hy}-{checkyear_hy}")
                                    # Values to the left of 1 have year - 1
                                    for month_hy in selected_period_hy[:index_of_1_hy]:
                                        new_list_hy.append(f"{month_hy}-{checkyear_hy - 1}")
                                
                            else:
                                for month_hy in selected_period_hy:
                                    new_list_hy.append(f"{month_hy}-{checkyear_hy}")
                                    
                        ordered_list_hy = []
                        for month_z1 in selected_period_hy:
                            # Find the corresponding year-month combination for each month in selected_period_hy
                            for item_z1 in new_list_hy:
                                if item_z1.startswith(f"{month_z1}-"):
                                    ordered_list_hy.append(item_z1)
                                    break
                                
                        new_list_hy = ordered_list_hy                                                                                                
                                                                                                                                                                                                              
                        parameters3H = [slab_id,re_month,re_year,emp_gender,slab_type,gross_salary,slab_applicable_status]
                        cursor.callproc("stp_getSlabForEmployee",parameters3H)
                        for result23H in cursor.stored_results():
                            myH = list(result23H.fetchall())
                            pt_actualH = myH[0][0]
                            statuscheckH = myH[0][1]
                            # IF SLAB EXISTS THEN SENDING yes ELSE noslabdefined
                            if (statuscheckH == "yes"):
                                                                                
                                # GET START AND END PERIOD LOGIC                                                
                                last_period_hy = new_list_hy[-1]
                                first_period_hy = new_list_hy[0]                                                
                                
                                salary_period_hy = (first_period_hy + " to " + last_period_hy)
                                salary_period_out = salary_period_hy
                                pt_actual_out = pt_actualH

                            if (statuscheckH == "noslabdefined"):                                            
                                salary_period_out = None                                
                                pt_actual_out = 0

                    if slab_freq == "Yearly":
                        checkmonth_y=int(re_month)                                     
                        checkyear_y=int(re_year)
                        matching_period_y = None
                        new_list_y = []

                        parametersY = [slab_id]
                        cursor.callproc("stp_getperiodforYearly",parametersY)
                        for result23Y in cursor.stored_results():
                            myY = list(result23Y.fetchall())
                            periodSE = myY[0][0] if myY[0][0] is not None else None
                                                                                                                                                                                                    
                            
                        if periodSE:                                           
                            slab_months_listY = slab_months_challan.split(',')
                            slab_months_listY = list(map(int, slab_months_listY))
                            
                            start_monthYY, end_monthYY = map(int, periodSE.split('-'))                                            
                            for monthYY in slab_months_listY:
                                if checkmonth_y < start_monthYY:
                                    # Assign checkyear_y to months less than basemonth, year - 1 to basemonth and months greater than it
                                    if monthYY < start_monthYY:
                                        new_list_y.append(f"{monthYY}-{checkyear_y}")
                                    else:
                                        new_list_y.append(f"{monthYY}-{checkyear_y - 1}")
                                else:
                                    # Assign the same year to basemonth and months greater than it, year +1 to months less than it
                                    if monthYY >= start_monthYY:
                                        new_list_y.append(f"{monthYY}-{checkyear_y}")
                                    else:
                                        new_list_y.append(f"{monthYY}-{checkyear_y + 1}")                                                                                          

                            # Initialize startperiodz and endperiodz to None
                            startperiodz = None
                            endperiodz = None

                            # Iterate through new_list_y to find the matching values
                            for period_y1 in new_list_y:
                                month_y1, year_y1 = map(int, period_y1.split('-'))
                                
                                # Check if the month matches the start period
                                if month_y1 == start_monthYY:
                                    startperiodz = period_y1
                                
                                # Check if the month matches the end period
                                if month_y1 == end_monthYY:
                                    endperiodz = period_y1 
                            
                                            
                            parameters3Y = [slab_id,re_month,re_year,emp_gender,slab_type,gross_salary,slab_applicable_status]
                            cursor.callproc("stp_getSlabForEmployee",parameters3Y)
                            for result23YY in cursor.stored_results():
                                myYY = list(result23YY.fetchall())
                                pt_actualYY = myYY[0][0]
                                statuscheckYY = myYY[0][1]
                        # IF SLAB EXISTS THEN SENDING yes ELSE noslabdefined 
                                if (statuscheckYY == "yes"):
                                                        
                            # GET START AND END PERIOD LOGIC                                                        
                                                                                  
                                    salary_period_y = (startperiodz + " to " + endperiodz)
                                    salary_period_out = salary_period_y
                                    pt_actual_out = pt_actualYY 
                                                                       
                                if (statuscheckYY == "noslabdefined"):                                                    
                                    salary_period_out = None                                
                                    pt_actual_out = 0                                                                                                  
                        
                        else:
                            salary_period_out = None                                
                            pt_actual_out = 0
                                                                                                                                                                                                                                                                
    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])         
        print("error-" + e)
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        return comp_id_out, state_id_out, city_id_out, act_id_out, site_id_out, re_year_out, re_month_out, emp_code_out, gross_salary_out, pt_actual_out, salary_period_out,slab_freq
    

def lwf_calculation(user_session,comp_id_in, re_year_in, re_month_in, act_id_id_in,emp_gender_in ,emp_code_in, gross_salary_in, city_id_id_in, state_id_id_in, site_id_in,slab_freq,slab_period,slab_id,slab_months):
    Db.closeConnection()
    m = Db.get_connection()
    cursor = m.cursor()

    try:
        
        year = re_year_in
        month = re_month_in
        LWF1 = 2
        
        emp_code = emp_code_in,
        branch_id_id = site_id_in,
        city_id_id = city_id_id_in,
        # gross_sal = gross_salary_in,
        gross_sal = float(gross_salary_in)


        comp_id_out = comp_id_in,
        state_id_out = state_id_id_in,
        city_id_out = city_id_id_in,
        act_id_out = act_id_id_in,     
        site_id_out = site_id_in ,

         
        slab_freq = slab_freq
        slab_months = slab_months
        slab_id = slab_id
        slab_period = slab_period
        slab_months_list = slab_months.replace(" ", "").split(",")                   
                            
        if slab_freq == 'Monthly':
                                        
                month_int_str = str(int(month))
                currentMonthAndYear = f"{month_int_str}-{year}"
                current_date = datetime.strptime(currentMonthAndYear, '%m-%Y')

                slab_months_list = slab_months.replace(" ", "").split(",")

                new_list_y = []
                checkmonth_y = int(month)
                checkyear_y = int(year)
                                                            
                start_monthYY, end_monthYY = map(int, slab_period.split('-'))

                for monthYY in slab_months_list:
                    
                    monthYY_int = int(monthYY)

                    if checkmonth_y < start_monthYY:
                        if monthYY_int < start_monthYY:
                            new_list_y.append(f"{monthYY_int}-{checkyear_y}")
                        else:
                            new_list_y.append(f"{monthYY_int}-{checkyear_y - 1}")
                    else:
                        if monthYY_int >= start_monthYY:
                            new_list_y.append(f"{monthYY_int}-{checkyear_y}")
                        else:
                            new_list_y.append(f"{monthYY_int}-{checkyear_y + 1}")
                                                            
                if new_list_y:
                    date_objects = [datetime.strptime(item, "%m-%Y") for item in new_list_y]
                    
                    date_objects.sort()

                    start_period = None
                    end_period = None

                    for date in date_objects:
                        if date < current_date:
                            start_period = date
                        elif date >= current_date:
                            end_period = date  
                            break 

                    if start_period and end_period:
                        start_month_incremented = start_period.month + 1
                        start_year = start_period.year

                        if start_month_incremented > 12:
                            start_month_incremented = 1
                            start_year += 1

                        challan_period = f"{start_month_incremented}-{start_year} to {end_period.month}-{end_period.year}"                                                       
                                                            
                def parse_date(date_str):
                    return datetime.strptime(date_str, '%m-%Y')

                dates = [parse_date(date_str) for date_str in new_list_y]
                future_dates = [date for date in dates if date >= current_date]
                    
                if future_dates:
                    next_closest_date = min(future_dates, key=lambda x: abs((x - current_date).days))
                    next_closest_date_str = next_closest_date.strftime('%m-%Y')
                else:
                    next_closest_date_str = None
                    
                if next_closest_date_str:
                    month_str, year_str = next_closest_date_str.split('-')
                    challan_month = int(month_str)
                    challan_year = int(year_str)
                    
                    if len(slab_months_list) == 2:
                                                        
                        next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                        new_date = next_closest_date - relativedelta(months=5)
                        new_date_str = new_date.strftime('%m-%Y')
                        # challan_period = f"{new_date_str} to {next_closest_date_str}"
                        new_date_str = f"{new_date.month}-{new_date.year}"
                        next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                        challan_period = f"{new_date_str} to {next_closest_date_str}"
                                                                            
                        if month_int_str in slab_months_list:
                            
                            try:
                                cursor.callproc(
                                    "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                                    [slab_id, gross_sal, month, year, challan_period]
                                )
                            except Exception as e:
                                print(f"Error: {e}")
                            
                            for result in cursor.stored_results():
                                result_set = result.fetchall()
                                if result_set:
                                    result_value = result_set[0][0]
                                    
                                    if result_value == 0:
                                        print("No Slab")
                                        description = f"Employee Code - {emp_code} : Slab Not Applicable"
                                    else:
                                        setEmployee = result_set[0][1]
                                        setEmployer = result_set[0][2]
                                        challan_period1 = result_set[0][3]
                    
                    elif len(slab_months_list) == 4:
                                                        
                        next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                        new_date = next_closest_date - relativedelta(months=2)
                        new_date_str = new_date.strftime('%m-%Y')
                        # challan_period = f"{new_date_str} to {next_closest_date_str}"
                        new_date_str = f"{new_date.month}-{new_date.year}"
                        next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                        challan_period = f"{new_date_str} to {next_closest_date_str}"
                                                                            
                        if month_int_str in slab_months_list:
                            
                            try:
                                cursor.callproc(
                                    "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                                    [slab_id, gross_sal, month, year, challan_period]
                                )
                            except Exception as e:
                                print(f"Error: {e}")

                            
                            for result in cursor.stored_results():
                                result_set = result.fetchall()
                                if result_set:
                                    result_value = result_set[0][0]
                                    
                                    if result_value == 0:
                                        print("No Slab")
                                        description = f"Employee Code - {emp_code} : Slab Not Applicable"
                                    else:
                                        setEmployee = result_set[0][1]
                                        setEmployer = result_set[0][2]
                                        challan_period1 = result_set[0][3]
                                        
                    elif len(slab_months_list) == 1:
                                                        
                        next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                        new_date = next_closest_date - relativedelta(months=11)
                        new_date_str = new_date.strftime('%m-%Y')
                        # challan_period = f"{new_date_str} to {next_closest_date_str}"
                        new_date_str = f"{new_date.month}-{new_date.year}"
                        next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                        challan_period = f"{new_date_str} to {next_closest_date_str}"
                                                                            
                        if month_int_str  in slab_months_list:

                            cursor.callproc(
                                "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                                [
                                    slab_id, gross_sal, month, year, challan_period
                                ]
                            )
                            for result in cursor.stored_results():
                                result_set = result.fetchall()
                                if result_set:
                                    result_value = result_set[0][0]
                                    
                                    if result_value == 0:
                                        print("No Slab")
                                        description = f"Employee Code - {emp_code} : Slab Not Applicable"
                                    else:
                                        setEmployee = result_set[0][1]
                                        setEmployer = result_set[0][2]
                                        challan_period1 = result_set[0][3]
                    
                    elif len(slab_months_list) == 12:
                                                        
                        next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                        new_date = next_closest_date
                        new_date_str = new_date.strftime('%m-%Y')
                        # challan_period = f"{new_date_str}"
                        challan_period = f"{new_date.month}-{new_date.year}"
                            
                        cursor.callproc(
                            "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                            [
                                slab_id, gross_sal, month, year, challan_period
                            ]
                        )
                        for result in cursor.stored_results():
                                result_set = result.fetchall()
                                if result_set:
                                    result_value = result_set[0][0]
                                    
                                    if result_value == 0:
                                        print("No Slab")
                                        description = f"Employee Code - {emp_code} : Slab Not Applicable"
                                    else:
                                        setEmployee = result_set[0][1]
                                        setEmployer = result_set[0][2]
                                        challan_period1 = result_set[0][3]
                   
        else: 
                                                
            month_int_str = str(int(month))
            currentMonthAndYear = f"{month_int_str}-{year}"
            current_date = datetime.strptime(currentMonthAndYear, '%m-%Y')
            slab_months_list = slab_months.replace(" ", "").split(",")
            slab_months_with_year = [f"{month}-{year}" for month in slab_months_list]
            def parse_date(date_str):
                return datetime.strptime(date_str, '%m-%Y')
            dates = [parse_date(date_str) for date_str in slab_months_with_year]
            future_dates = [date for date in dates if date >= current_date]
            if future_dates:
                next_closest_date = min(future_dates, key=lambda x: abs((x - current_date).days))
                next_closest_date_str = next_closest_date.strftime('%m-%Y')
            else:
                next_closest_date_str = None
            if next_closest_date_str:
                month_str, year_str = next_closest_date_str.split('-')
                challan_month = int(month_str)
                challan_year = int(year_str)
                
                if slab_freq == 'Yearly':
                                                    
                    next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                    new_date = next_closest_date - relativedelta(months=11)
                    new_date_str = new_date.strftime('%m-%Y')
                    # challan_period = f"{new_date_str} to {next_closest_date_str}"
                    new_date_str = f"{new_date.month}-{new_date.year}"
                    next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                    challan_period = f"{new_date_str} to {next_closest_date_str}"
                                            
                    if month_int_str in slab_months_list: 
                        location_dol_flag = 0
                        cursor.callproc(
                            "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                            [slab_id, gross_sal, month, year, challan_period]
                        )
                        for result in cursor.stored_results():
                            result_set = result.fetchall()
                            if result_set:
                                result_value = result_set[0][0]
                                
                                if result_value == 0:
                                    print("No Slab")
                                else:
                                    setEmployee = result_set[0][1]
                                    setEmployer = result_set[0][2]
                                    challan_period1 = result_set[0][3]
                elif slab_freq == 'Half Yearly':
                                                    
                    next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                    new_date = next_closest_date - relativedelta(months=5)
                    new_date_str = new_date.strftime('%m-%Y')
                    # challan_period = f"{new_date_str} to {next_closest_date_str}"
                    new_date_str = f"{new_date.month}-{new_date.year}"
                    next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                    challan_period = f"{new_date_str} to {next_closest_date_str}"
                    if month_int_str in slab_months_list: 
                        location_dol_flag = 0
                        cursor.callproc(
                            "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                            [slab_id, gross_sal, month, year, challan_period]
                        )
                        for result in cursor.stored_results():
                            result_set = result.fetchall()
                            if result_set:
                                result_value = result_set[0][0]
                                
                                if result_value == 0:
                                    print("No Slab")
                                else:
                                    setEmployee = result_set[0][1]
                                    setEmployer = result_set[0][2]
                                    challan_period1 = result_set[0][3]
                
                elif slab_freq == 'Quarterly':
                                                    
                    next_closest_date = datetime.strptime(next_closest_date_str, '%m-%Y')
                    new_date = next_closest_date - relativedelta(months=2)
                    new_date_str = new_date.strftime('%m-%Y')
                    # challan_period = f"{new_date_str} to {next_closest_date_str}"
                    new_date_str = f"{new_date.month}-{new_date.year}"
                    next_closest_date_str = f"{next_closest_date.month}-{next_closest_date.year}"
                    challan_period = f"{new_date_str} to {next_closest_date_str}"
                    if month_int_str in slab_months_list: 
                        location_dol_flag = 0
                        cursor.callproc(
                            "stp_getDetailsOfSlabForLWFFromTbl_slab_PSN",
                            [slab_id, gross_sal, month, year, challan_period]
                        )
                        for result in cursor.stored_results():
                            result_set = result.fetchall()
                            if result_set:
                                result_value = result_set[0][0]
                                
                                if result_value == 0:
                                    print("No Slab")
                                else:
                                    setEmployee = result_set[0][1]
                                    setEmployer = result_set[0][2]
                                    challan_period1 = result_set[0][3]
                                       
            # response_data = {"status": 10}
            # return JsonResponse(response_data)
            
            # cursor.close()
            # m.close()

        # return JsonResponse({"status": "no discrepancies found"})


    except Exception as e:
        tb = traceback.extract_tb(e.__traceback__)
        fun = tb[0].name
        cursor.callproc("stp_error_log",[fun,str(e),user])         
        print("error-" + e)
    finally:
        cursor.close()
        m.commit()
        m.close()
        Db.closeConnection()
        return comp_id_out, state_id_out, city_id_out, act_id_out, site_id_out,year, month, emp_code, gross_sal, setEmployee,setEmployer,challan_period1

           


